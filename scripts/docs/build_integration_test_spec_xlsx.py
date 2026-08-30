#!/usr/bin/env python3
"""結合テスト仕様書(Excel)生成スクリプト。
ケイスタンプ株式会社が、実装済みの結合テスト(API統合・RLS)と機能連携を整理し、
docs/test/AI人材採用マッチングシステム_結合テスト仕様書.xlsx を生成する(である調)。
記載はいずれも実装・自動テストの事実に基づき、未実装の挙動は記載しない。
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = "docs/test/AI人材採用マッチングシステム_結合テスト仕様書.xlsx"
FONT = "Yu Gothic"
TITLE_FONT = Font(name=FONT, bold=True, size=14, color="1F3864")
HEAD_FONT = Font(name=FONT, bold=True, size=10, color="FFFFFF")
BODY_FONT = Font(name=FONT, size=10)
KEY_FONT = Font(name=FONT, bold=True, size=10, color="1F3864")
HEAD_FILL = PatternFill("solid", fgColor="1F3864")
KEY_FILL = PatternFill("solid", fgColor="E8ECF6")
OK_FILL = PatternFill("solid", fgColor="E7F4E4")
_thin = Side(style="thin", color="C6CEDD")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
WRAP = Alignment(wrap_text=True, vertical="top")
HEAD_AL = Alignment(wrap_text=True, vertical="center", horizontal="center")


def add(wb, name, title, headers, rows, widths=None, ok_col=None):
    ws = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    for j, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=j, value=h)
        c.font, c.fill, c.alignment, c.border = HEAD_FONT, HEAD_FILL, HEAD_AL, BORDER
    for i, row in enumerate(rows, 4):
        for j, val in enumerate(row, 1):
            c = ws.cell(row=i, column=j, value=val)
            c.font, c.alignment, c.border = BODY_FONT, WRAP, BORDER
            if ok_col and j == ok_col and val == "OK":
                c.fill = OK_FILL
                c.alignment = Alignment(wrap_text=True, vertical="top", horizontal="center")
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:{get_column_letter(len(headers))}{3 + len(rows)}"
    for j in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(j)].width = (
            widths[j - 1] if widths and j - 1 < len(widths) else 22
        )
    ws.row_dimensions[1].height = 22
    return ws


def cover(wb, name, title, pairs):
    ws = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    for i, (k, v) in enumerate(pairs, 3):
        kc = ws.cell(row=i, column=1, value=k)
        kc.font, kc.fill, kc.alignment, kc.border = KEY_FONT, KEY_FILL, WRAP, BORDER
        vc = ws.cell(row=i, column=2, value=v)
        vc.font, vc.alignment, vc.border = BODY_FONT, WRAP, BORDER
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 94
    return ws


wb = Workbook()
wb.remove(wb.active)

# ─── 00_表紙 ───
cover(wb, "00_表紙", "AI人材採用マッチングシステム 結合テスト仕様書", [
    ("文書名", "AI人材採用マッチングシステム 結合テスト仕様書(Excel形式)"),
    ("対象システム", "AI人材採用マッチングシステム"),
    ("テスト工程", "結合テスト(Integration Test)"),
    ("版数", "1.0"),
    ("作成日", "2026-06-20"),
    ("作成者", "ケイスタンプ株式会社"),
    ("対象バージョン", "0.1.0 / branch: feat/ai-recruitment-mvp / commit: 1a46f42"),
    ("対象範囲",
     "FR-01〜FR-10 の機能間連携。API を起点に、DB・RLS・認証・AI分析/マッチング・ジョブ・"
     "メッセージ/通知・採用フロー・多言語・非機能(レート制限/契約/死活)の連携を検証する。"),
    ("テスト方式",
     "自動化結合テスト。Fastify を実起動相当(in-process)で起動し、実DB相当の PostgreSQL"
     "(PGlite + pgvector, RLS有効)に対して検証する。AI・ウイルススキャン・ストレージは決定論モックを用いる。"),
    ("前提・制約",
     "本書はケイスタンプ株式会社が作成した結合テストの仕様書である。各テストケースは自動結合テスト"
     "(統合テスト/RLSテスト)に対応する。記載は実装と自動テストの事実に基づき、未実装の挙動は記載しない。"),
])

# ─── 01_改訂履歴 ───
add(wb, "01_改訂履歴", "01_改訂履歴",
    ["版数", "日付", "変更区分", "変更内容", "作成者", "レビュー状態"],
    [["1.0", "2026-06-20", "新規作成",
      "初版。実装済みの結合テスト(統合78件・RLS26件)と機能連携を整理して作成した。",
      "ケイスタンプ株式会社", "未レビュー(要レビュー)"]],
    [8, 12, 12, 64, 18, 16])

# ─── 02_テスト方針 ───
add(wb, "02_テスト方針", "02_テスト方針",
    ["区分", "内容"],
    [
        ["目的", "機能間の連携が要求どおりに動作することを確認する。"],
        ["対象工程", "結合テストである。単体テスト・E2E・性能・アクセシビリティは別工程とする。"],
        ["テスト対象", "API を起点とした機能横断の連携(API↔DB↔RLS↔認証↔AI↔ジョブ↔通知)である。"],
        ["テスト方式",
         "自動化する。Fastify を in-process で起動し、実DB相当の PostgreSQL(PGlite+pgvector, RLS有効)を用いる。"
         "AI・ウイルススキャン・ストレージは決定論モックとし、結果を再現可能にする。"],
        ["合否基準", "期待結果と、実際の応答・DB状態・権限挙動が一致すればOK、不一致はNGとする。"],
        ["進入基準", "対象機能の単体テストが完了していることとする。"],
        ["退出基準", "全結合テストケースがOKであることとする。"],
        ["自動テストとの対応",
         "各ケースは vitest の統合テスト(apps/api/tests/*.integration.test.ts)及びRLSテスト"
         "(apps/api/tests/*.rls.test.ts)に対応する(根拠列を参照)。"],
        ["実施方法", "`npm run test:integration`(統合)及び `npm run test:rls`(RLS)で実行する。"],
        ["最新結果", "2026-06-20 時点で、統合78件・RLS26件の全件が成功している(commit 1a46f42)。"],
        ["確認者", "ケイスタンプ株式会社"],
    ],
    [16, 92])

# ─── 03_テスト環境 ───
add(wb, "03_テスト環境", "03_テスト環境",
    ["項目", "内容"],
    [
        ["実行環境", "Node.js 20 LTS"],
        ["ランタイム", "ARS_RUNTIME=local(PGlite + pgvector、決定論モック)。本番は supabase ランタイム。"],
        ["データベース", "PostgreSQL 互換(pgvector 有効、全テーブル RLS 有効)"],
        ["認証", "ローカルJWT(本番は GoTrue 相当)。リクエスト毎に RLS コンテキストを設定する。"],
        ["AIプロバイダ", "モック(決定論)。本番は外部LLM(Anthropic / OpenAI、設定で切替)。"],
        ["ウイルススキャン", "モック(EICAR で検知を模擬)。本番は ClamAV。"],
        ["ストレージ", "ローカルファイルシステム(本番はオブジェクトストレージ)"],
        ["実行コマンド", "npm run test:integration / npm run test:rls"],
        ["テストデータ", "各ケースで初期化し、ケース間で独立させる。"],
        ["トレース", "全応答に x-correlation-id を付与する。"],
    ],
    [18, 90])

# ─── 04_テスト観点一覧 ───
add(wb, "04_テスト観点一覧", "04_テスト観点一覧",
    ["観点ID", "テスト観点", "関連FR", "主な連携点", "確認内容", "対応テスト"],
    [
        ["KP-01", "認証連携", "FR-01", "API↔DB(profiles/candidates)↔JWT", "登録・ログイン・確認・更新・認可", "auth.integration.test.ts"],
        ["KP-02", "ファイル処理連携", "FR-02", "API↔検証↔ウイルス↔ストレージ↔ジョブ", "アップロード検証・スキャン・解析起動", "resume / security-files.integration.test.ts"],
        ["KP-03", "AI分析連携", "FR-03", "API↔ジョブ↔LLM(モック)↔DB", "分析生成・取得・ロケール", "matching.integration.test.ts"],
        ["KP-04", "企業/求人連携", "FR-04/07", "API↔DB↔RLS", "企業・求人CRUD・公開範囲・検索", "company-job.integration.test.ts"],
        ["KP-05", "マッチング連携", "FR-05", "API↔pgvector↔スコアリング↔DB", "推薦・候補者ランキング・再現性", "matching.integration.test.ts"],
        ["KP-06", "人材検索連携", "FR-06", "API↔RLS(可視性)↔DB", "検索・絞り込み・詳細開示制御・比較", "talent.integration.test.ts"],
        ["KP-07", "メッセージ/通知連携", "FR-08", "API↔RLS↔通知", "会話・送信・重複排除・通知", "messaging.integration.test.ts"],
        ["KP-08", "採用フロー連携", "FR-10", "API↔RLS↔履歴↔通知", "応募・段階遷移・面接・ショートリスト", "recruitment.integration.test.ts"],
        ["KP-09", "多言語連携", "FR-09", "API(messageKey)↔クライアント辞書", "エラーの言語非依存キー返却", "messaging / auth(messageKey)"],
        ["KP-10", "認可境界(RLS)", "全FR", "SET ROLE↔ポリシー↔定義関数", "テナント分離・IDOR・否定系", "rls / security-negative.rls.test.ts"],
        ["KP-11", "非機能連携", "横断", "API横断", "レート制限429・OpenAPI契約・死活", "ratelimit / openapi / health.integration.test.ts"],
        ["KP-12", "冪等性/競合", "FR-02/10", "並行書込↔unique/冪等キー", "二重応募・並行投入の整合", "idempotency.integration.test.ts"],
    ],
    [8, 18, 12, 30, 30, 34])

# ─── 05_結合テストケース ───
HDR = ["TC番号", "大項目", "中項目", "関連FR", "観点", "前提条件", "テスト手順", "入力データ",
       "期待結果", "確認API/画面", "区分", "根拠テスト", "判定", "実施日"]
D = "2026-06-20"
TT = "auth.integration.test.ts"
RS = "resume.integration.test.ts"
SF = "security-files.integration.test.ts"
MT = "matching.integration.test.ts"
CJ = "company-job.integration.test.ts"
TL = "talent.integration.test.ts"
MS = "messaging.integration.test.ts"
RC = "recruitment.integration.test.ts"
RL = "ratelimit.integration.test.ts"
ID = "idempotency.integration.test.ts"
OA = "openapi.integration.test.ts"
HL = "health.integration.test.ts"
RZ = "rls.rls.test.ts"
RN = "security-negative.rls.test.ts"

CASES = [
    # 認証 FR-01
    ["TC-AUTH-01", "認証", "ユーザー登録(求職者)", "FR-01", "KP-01", "未登録である",
     "登録APIを呼ぶ", "email/password/role=job_seeker/displayName",
     "201。profiles と candidates が作成され、Session を返す。", "POST /api/auth/register", "正常", TT, "OK", D],
    ["TC-AUTH-02", "認証", "ユーザー登録(企業)", "FR-01", "KP-01", "未登録である",
     "登録APIを呼ぶ", "role=company_member",
     "201。profiles が作成され、Session を返す。", "POST /api/auth/register", "正常", TT, "OK", D],
    ["TC-AUTH-03", "認証", "重複メール登録", "FR-01", "KP-01", "同一メールが登録済である",
     "同一メールで再登録する", "既存email",
     "409(CONFLICT)。messageKey は auth.taken。", "POST /api/auth/register", "異常", TT, "OK", D],
    ["TC-AUTH-04", "認証", "登録バリデーション", "FR-01", "KP-01", "—",
     "不正値で登録する", "8文字未満のpassword",
     "422(VALIDATION)。項目エラーを返す。", "POST /api/auth/register", "異常", TT, "OK", D],
    ["TC-AUTH-05", "認証", "ログイン成功", "FR-01", "KP-01", "登録済である",
     "正資格情報でログインする", "正email/password",
     "200。accessToken と user を返す。", "POST /api/auth/login", "正常", TT, "OK", D],
    ["TC-AUTH-06", "認証", "ログイン失敗", "FR-01", "KP-01", "登録済である",
     "誤パスワードでログインする", "誤password",
     "401(UNAUTHORIZED)。messageKey は auth.invalidCredentials。", "POST /api/auth/login", "異常", TT, "OK", D],
    ["TC-AUTH-07", "認証", "メール確認", "FR-01", "KP-01", "確認トークンを保持する",
     "確認APIを呼ぶ", "verifyトークン",
     "200。emailVerified が true となる。", "POST /api/auth/verify-email", "正常", TT, "OK", D],
    ["TC-AUTH-08", "認証", "自プロフィール取得", "FR-01", "KP-01", "認証済である",
     "認証付きで自情報を取得する", "Bearerトークン",
     "200。AuthUser を返す。", "GET /api/auth/me", "正常", TT, "OK", D],
    ["TC-AUTH-09", "認証", "未認証アクセス", "FR-01", "KP-10", "未認証である",
     "トークンなしで保護APIを呼ぶ", "—",
     "401(UNAUTHORIZED)。", "GET /api/auth/me", "異常", TT, "OK", D],
    ["TC-AUTH-10", "認証", "アカウント更新", "FR-01", "KP-01", "認証済である",
     "表示名・言語・パスワードを更新する", "displayName/locale/password",
     "200。profiles に反映される。", "PATCH /api/auth/account", "正常", TT, "OK", D],
    # 履歴書 FR-02
    ["TC-RES-01", "履歴書", "正常アップロード", "FR-02", "KP-02", "求職者で認証済である",
     "PDFをアップロードする", "正PDF(≤10MB)",
     "201。解析ジョブが起動し、inlineで succeeded となる。", "POST /api/candidates/me/resume", "正常", RS, "OK", D],
    ["TC-RES-02", "履歴書", "サイズ超過", "FR-02", "KP-02", "求職者で認証済である",
     "10MB超のファイルを送る", "11MBファイル",
     "413(PAYLOAD_TOO_LARGE)。resume.upload.tooLarge。", "POST /api/candidates/me/resume", "異常", RS, "OK", D],
    ["TC-RES-03", "履歴書", "不正拡張子/MIME", "FR-02", "KP-02", "求職者で認証済である",
     "非対応形式を送る", ".exe等",
     "415。badExtension / badMime を返す。", "POST /api/candidates/me/resume", "異常", RS, "OK", D],
    ["TC-RES-04", "履歴書", "内容不一致(マジックバイト)", "FR-02", "KP-02", "求職者で認証済である",
     "拡張子と中身が不整合のファイルを送る", "PDF拡張子だが中身が非PDF",
     "415。resume.upload.badContent。", "POST /api/candidates/me/resume", "異常", SF, "OK", D],
    ["TC-RES-05", "履歴書", "ウイルス検知", "FR-02", "KP-02", "求職者で認証済である",
     "EICARを含むファイルを送る", "EICAR検体",
     "422(VIRUS_DETECTED)。保存せず resume.upload.infected を返す。", "POST /api/candidates/me/resume", "異常", SF, "OK", D],
    ["TC-RES-06", "履歴書", "解析ジョブ状態取得", "FR-02", "KP-02", "アップロード済である",
     "解析ジョブ状態を取得する", "parseJobId",
     "200。status(pending/processing/succeeded/failed)を返す。", "GET …/parse-jobs/:id", "正常", RS, "OK", D],
    ["TC-RES-07", "履歴書", "解析リトライ", "FR-02", "KP-02", "解析が失敗している",
     "リトライAPIを呼ぶ", "parseJobId",
     "再enqueueされ、再処理される。", "POST …/parse-jobs/:id/retry", "正常", RS, "OK", D],
    ["TC-RES-08", "履歴書", "他者履歴書のDL不可", "FR-02", "KP-10", "応募/保存の関係がない",
     "企業が無関係候補者の履歴書をDLする", "他候補者のresumeId",
     "403/404。RLS と API 認可で拒否する。", "GET /api/resumes/:id/download", "異常", SF, "OK", D],
    # AI分析 FR-03
    ["TC-ANL-01", "AI分析", "分析生成", "FR-03", "KP-03", "履歴書解析が完了している",
     "分析生成APIを呼ぶ", "locale",
     "200。skill_analyses が生成され、スキル/年数/提案を返す。", "POST /api/candidates/me/analysis", "正常", MT, "OK", D],
    ["TC-ANL-02", "AI分析", "分析取得", "FR-03", "KP-03", "分析が存在する",
     "分析取得APIを呼ぶ", "—",
     "200。最新の SkillAnalysis を返す。", "GET /api/candidates/me/analysis", "正常", MT, "OK", D],
    ["TC-ANL-03", "AI分析", "未分析時の取得", "FR-03", "KP-03", "分析が存在しない",
     "分析取得APIを呼ぶ", "—",
     "404。messageKey は analysis.none。", "GET /api/candidates/me/analysis", "異常", MT, "OK", D],
    ["TC-ANL-04", "AI分析", "ロケール指定生成", "FR-03/09", "KP-03", "履歴書解析が完了している",
     "locale を指定して生成する", "locale=en",
     "指定言語で分析が生成される。", "POST /api/candidates/me/analysis", "正常", MT, "OK", D],
    # 企業/求人 FR-04/07
    ["TC-CMP-01", "企業/求人", "企業作成(企業)", "FR-04", "KP-04", "company_memberで認証済である",
     "企業作成APIを呼ぶ", "name/industry/size",
     "201。company と company_members が作成される。", "POST /api/companies", "正常", CJ, "OK", D],
    ["TC-CMP-02", "企業/求人", "企業作成(権限なし)", "FR-04", "KP-10", "job_seekerで認証済である",
     "企業作成APIを呼ぶ", "—",
     "403(FORBIDDEN)。role により拒否する。", "POST /api/companies", "異常", CJ, "OK", D],
    ["TC-CMP-03", "企業/求人", "公開企業検索", "FR-07", "KP-04", "—",
     "企業検索を実行する", "q/industry/size",
     "200。公開企業一覧と openJobCount を返す。", "GET /api/companies", "正常", CJ, "OK", D],
    ["TC-JOB-01", "企業/求人", "求人作成", "FR-04", "KP-04", "企業メンバーである",
     "求人作成APIを呼ぶ", "title/skills/salary/visibility=public/status=open",
     "201。job と job_skills、埋め込みが作られる。", "POST /api/companies/:id/jobs", "正常", CJ, "OK", D],
    ["TC-JOB-02", "企業/求人", "公開求人の閲覧", "FR-07", "KP-04", "open+publicの求人がある",
     "一般ユーザーで求人一覧を取得する", "—",
     "公開求人が一覧に表示される。", "GET /api/jobs", "正常", CJ, "OK", D],
    ["TC-JOB-03", "企業/求人", "非公開求人の秘匿", "FR-07", "KP-10", "draft/privateの求人がある",
     "未認証/他社で一覧取得する", "—",
     "非公開求人は一覧に表示されない(RLS)。", "GET /api/jobs", "異常", CJ, "OK", D],
    ["TC-JOB-04", "企業/求人", "他社求人の編集不可", "FR-04", "KP-10", "他社の求人である",
     "他社メンバーが求人を更新する", "jobId",
     "403。RLS とテナント分離で拒否する。", "PATCH /api/jobs/:id", "異常", RN, "OK", D],
    ["TC-JOB-05", "企業/求人", "求人検索フィルタ", "FR-07", "KP-04", "複数求人がある",
     "条件付きで検索する", "q/workStyle/skills",
     "条件に一致する求人のみ返る。", "GET /api/jobs", "正常", CJ, "OK", D],
    # マッチング FR-05
    ["TC-MAT-01", "マッチング", "推薦取得(求職者)", "FR-05", "KP-05", "分析済の求職者と求人がある",
     "推薦APIを呼ぶ", "—",
     "200。score 降順で reason 付きの推薦を返す。", "GET /api/candidates/me/recommendations", "正常", MT, "OK", D],
    ["TC-MAT-02", "マッチング", "候補者ランキング(求人)", "FR-05", "KP-05", "求人と候補者がある",
     "求人の候補者一覧を取得する", "jobId",
     "200。score 付きの候補者を返す。", "GET /api/jobs/:id/candidates", "正常", MT, "OK", D],
    ["TC-MAT-03", "マッチング", "スコア再現性", "FR-05", "KP-05", "同一条件である",
     "同条件で2回算定する", "同一入力",
     "同一スコアを返す(決定論・版管理 match-v1)。", "(scoring)", "正常", MT, "OK", D],
    ["TC-MAT-04", "マッチング", "一致/不足スキル", "FR-05", "KP-05", "求人と候補者がある",
     "マッチ結果を確認する", "—",
     "matched_skills / missing_skills が算出される。", "GET …/recommendations", "正常", MT, "OK", D],
    # 人材検索 FR-06
    ["TC-TAL-01", "人材検索", "候補者検索", "FR-06", "KP-06", "company_memberで認証済である",
     "人材検索を実行する", "—",
     "200。可視範囲の候補者一覧を返す。", "GET /api/talent", "正常", TL, "OK", D],
    ["TC-TAL-02", "人材検索", "経験年数フィルタ", "FR-06", "KP-06", "候補者が複数いる",
     "minYears を指定する", "minYears=3",
     "条件以上の候補者のみ返る。", "GET /api/talent", "正常", TL, "OK", D],
    ["TC-TAL-03", "人材検索", "スキルフィルタ", "FR-06", "KP-06", "候補者が複数いる",
     "skills を指定する", "skills=React",
     "条件に合う候補者のみ返る。", "GET /api/talent", "正常", TL, "OK", D],
    ["TC-TAL-04", "人材検索", "詳細の開示制御", "FR-06", "KP-10", "応募/保存の関係有無が異なる",
     "候補者詳細を取得する", "candidateId",
     "連絡先・履歴書は関係成立時のみ開示される。", "GET /api/talent/:id", "正常/異常", TL, "OK", D],
    ["TC-TAL-05", "人材検索", "候補者比較", "FR-06/10", "KP-06", "対象候補者がいる",
     "2〜5名の比較を要求する", "candidateIds/jobId",
     "200。比較結果(スキル/スコア)を返す。", "POST /api/compare", "正常", RC, "OK", D],
    # メッセージ/通知 FR-08
    ["TC-MSG-01", "メッセージ", "会話作成+初回送信", "FR-08", "KP-07", "二者が存在する",
     "会話を作成し初回メッセージを送る", "participant/initialMessage",
     "会話・メッセージが作成され、相手に通知される。", "POST /api/conversations", "正常", MS, "OK", D],
    ["TC-MSG-02", "メッセージ", "メッセージ送信", "FR-08", "KP-07", "会話メンバーである",
     "メッセージを送信する", "body/clientToken",
     "メッセージが追加され、相手に通知される。", "POST /api/conversations/:id/messages", "正常", MS, "OK", D],
    ["TC-MSG-03", "メッセージ", "重複送信の抑止", "FR-08", "KP-12", "同一clientTokenを使う",
     "同一トークンで二重送信する", "同一clientToken",
     "メッセージは1件のみ作成される。", "POST …/messages", "正常", MS, "OK", D],
    ["TC-MSG-04", "メッセージ", "非メンバーのアクセス不可", "FR-08", "KP-10", "会話メンバーでない",
     "他人の会話を取得/送信する", "convId",
     "403。RLS(is_conversation_member)で拒否する。", "GET/POST …/conversations/:id", "異常", MS, "OK", D],
    ["TC-NOT-01", "通知", "通知一覧", "FR-08", "KP-07", "通知が存在する",
     "通知一覧を取得する", "—",
     "200。自分宛の通知のみ返る。", "GET /api/notifications", "正常", MS, "OK", D],
    ["TC-NOT-02", "通知", "既読化", "FR-08", "KP-07", "未読通知がある",
     "個別/一括で既読化する", "notificationId",
     "readAt が更新される。", "POST …/read, /read-all", "正常", MS, "OK", D],
    # 採用フロー FR-10
    ["TC-APP-01", "採用フロー", "応募", "FR-10", "KP-08", "公開求人がある求職者である",
     "求人へ応募する", "jobId/coverNote",
     "201。application が作成され、企業へ通知される。", "POST /api/applications", "正常", RC, "OK", D],
    ["TC-APP-02", "採用フロー", "二重応募の抑止", "FR-10", "KP-12", "同一求人へ応募済である",
     "同一求人へ再応募する", "同一jobId",
     "409。unique(job,candidate)で拒否する。", "POST /api/applications", "異常", RC, "OK", D],
    ["TC-APP-03", "採用フロー", "並行応募の整合", "FR-10", "KP-12", "未応募である",
     "同一求人へ同時に2件応募する", "同時2リクエスト",
     "[201,409]。論理的に1件のみ作成される。", "POST /api/applications", "異常", ID, "OK", D],
    ["TC-APP-04", "採用フロー", "段階遷移(企業)", "FR-10", "KP-08", "応募が存在する",
     "段階を遷移させる", "stage=screening 等",
     "段階が更新され、履歴が記録され、通知される。", "PATCH …/applications/:id/stage", "正常", RC, "OK", D],
    ["TC-APP-05", "採用フロー", "候補者辞退", "FR-10", "KP-08", "応募が存在する",
     "候補者が辞退する", "stage=withdrawn",
     "辞退が反映される。", "PATCH …/applications/:id/stage", "正常", RC, "OK", D],
    ["TC-SHL-01", "採用フロー", "ショートリスト操作", "FR-10", "KP-08", "企業メンバーである",
     "追加・一覧・削除する", "candidateId/jobId",
     "追加/一覧/削除が企業境界内で行える。", "POST/GET/DELETE /api/shortlists", "正常", RC, "OK", D],
    ["TC-INT-01", "採用フロー", "面接提案(企業)", "FR-10", "KP-08", "応募が存在する",
     "面接を提案する", "scheduledAt/mode",
     "interview が作成され、候補者へ通知される。", "POST …/applications/:id/interviews", "正常", RC, "OK", D],
    ["TC-INT-02", "採用フロー", "面接応答(候補者)", "FR-10", "KP-08", "面接提案がある",
     "確認/辞退する", "response=confirmed/declined",
     "面接状態が更新される。", "POST /api/interviews/:id/respond", "正常", RC, "OK", D],
    # 多言語 FR-09
    ["TC-I18N-01", "多言語", "エラーの言語非依存キー", "FR-09", "KP-09", "—",
     "エラーを発生させ応答を確認する", "不正入力",
     "エラーは安定した messageKey を返し、クライアントが各言語へ変換できる。", "各API", "正常", TT, "OK", D],
    # 横断/非機能
    ["TC-SEC-01", "認可境界", "他者データの不可視", "全FR", "KP-10", "他者の候補者がいる",
     "他者の候補者を参照する", "他candidateId",
     "RLS(candidate_readable)により取得できない。", "(RLS)", "異常", RZ, "OK", D],
    ["TC-SEC-02", "認可境界", "否定系(越権)", "全FR", "KP-10", "権限のない操作を行う",
     "非企業メンバーが求人を編集する 等", "—",
     "RLS により拒否される(否定系を網羅)。", "(RLS)", "異常", RN, "OK", D],
    ["TC-OPS-02", "非機能", "レート制限", "横断", "KP-11", "低い上限を設定する",
     "上限を超えて連続要求する", "RATE_LIMIT_MAX=2 で3要求",
     "429(RATE_LIMITED)。error.rateLimited を返す。", "(各API)", "異常", RL, "OK", D],
    ["TC-OPS-03", "非機能", "OpenAPI契約", "横断", "KP-11", "アプリ起動済である",
     "契約を取得し主要経路を確認する", "—",
     "200。openapi/info/paths を含み主要経路が定義される。", "GET /openapi.json", "正常", OA, "OK", D],
    ["TC-OPS-04", "非機能", "死活/受入", "横断", "KP-11", "アプリ起動済である",
     "死活・受入を確認する", "—",
     "/health は200、/ready はDB往復後に準備状態を返す。", "GET /health, /ready", "正常", HL, "OK", D],
]
add(wb, "05_結合テストケース", "05_結合テストケース", HDR, CASES,
    [11, 12, 20, 10, 8, 22, 26, 22, 34, 24, 9, 26, 7, 12], ok_col=13)

# ─── 06_トレーサビリティ ───
add(wb, "06_トレーサビリティ", "06_トレーサビリティ(FR↔テストケース)",
    ["関連FR", "機能", "テスト観点", "テストケース範囲", "根拠テストファイル", "判定"],
    [
        ["FR-01", "認証・ユーザー管理", "KP-01/10", "TC-AUTH-01〜10, TC-I18N-01", "auth.integration.test.ts", "OK"],
        ["FR-02", "履歴書アップロード・解析", "KP-02", "TC-RES-01〜08", "resume / security-files.integration.test.ts", "OK"],
        ["FR-03", "AIスキル分析", "KP-03", "TC-ANL-01〜04", "matching.integration.test.ts", "OK"],
        ["FR-04", "企業・求人管理", "KP-04", "TC-CMP-01/02, TC-JOB-01/04", "company-job.integration.test.ts", "OK"],
        ["FR-05", "AIマッチング", "KP-05", "TC-MAT-01〜04", "matching.integration.test.ts", "OK"],
        ["FR-06", "人材検索・候補者詳細", "KP-06", "TC-TAL-01〜05", "talent.integration.test.ts", "OK"],
        ["FR-07", "企業・求人検索", "KP-04", "TC-CMP-03, TC-JOB-02/03/05", "company-job.integration.test.ts", "OK"],
        ["FR-08", "メッセージ・通知", "KP-07", "TC-MSG-01〜04, TC-NOT-01/02", "messaging.integration.test.ts", "OK"],
        ["FR-09", "多言語対応", "KP-09", "TC-ANL-04, TC-I18N-01", "auth / messaging(messageKey)", "OK"],
        ["FR-10", "比較・面接・採用フロー", "KP-08/12", "TC-APP-01〜05, TC-SHL-01, TC-INT-01/02, TC-TAL-05", "recruitment / idempotency.integration.test.ts", "OK"],
        ["横断", "認可境界(RLS)", "KP-10", "TC-SEC-01/02, TC-JOB-03/04, TC-MSG-04", "rls / security-negative.rls.test.ts", "OK"],
        ["横断", "非機能(制限/契約/死活)", "KP-11", "TC-OPS-02〜04", "ratelimit / openapi / health.integration.test.ts", "OK"],
    ],
    [10, 24, 12, 34, 36, 8], ok_col=6)

# ─── 07_不具合管理 ───
add(wb, "07_不具合管理", "07_不具合管理",
    ["No.", "検出日", "関連TC", "重大度", "事象", "原因", "対応", "ステータス", "確認者"],
    [["—", "2026-06-20", "—", "—", "現時点で起票なし。全結合テストケースが成功している。",
      "—", "—", "クローズ", "ケイスタンプ株式会社"]],
    [6, 12, 14, 10, 36, 22, 22, 12, 18])

# ─── 08_結果サマリ ───
add(wb, "08_結果サマリ", "08_結果サマリ",
    ["テスト分類", "対象", "件数", "成功", "失敗", "未実施", "実行コマンド", "結果"],
    [
        ["結合(API統合)", "機能間連携(API↔DB↔AI↔通知 等)", "78", "78", "0", "0",
         "npm run test:integration", "成功"],
        ["結合(RLS/認可)", "テナント分離・IDOR・否定系", "26", "26", "0", "0",
         "npm run test:rls", "成功"],
        ["合計", "—", "104", "104", "0", "0", "—", "全件成功(2026-06-20, commit 1a46f42)"],
    ],
    [16, 36, 8, 8, 8, 8, 28, 30])

wb.save(OUT)
import os
print("SHEETS:", len(wb.sheetnames), wb.sheetnames)
print("CASES:", len(CASES))
print(f"SAVED: {OUT} ({os.path.getsize(OUT)} bytes)")
