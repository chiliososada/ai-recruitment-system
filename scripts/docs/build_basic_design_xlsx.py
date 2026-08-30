#!/usr/bin/env python3
"""
基本設計書(Excel)生成スクリプト。
ケイスタンプ株式会社が、実装内容と要求仕様(FR-01〜FR-10)を突合して基本設計書ワークブックを
docs/design/basic/AI人材採用マッチングシステム_基本設計書.xlsx に生成する(である調)。
要求仕様書の細目と未突合の事項は「要確認」として要件担当と確認する。
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = "docs/design/basic/AI人材採用マッチングシステム_基本設計書.xlsx"
FONT = "Yu Gothic"
TITLE_FONT = Font(name=FONT, bold=True, size=14, color="1F3864")
HEAD_FONT = Font(name=FONT, bold=True, size=10, color="FFFFFF")
BODY_FONT = Font(name=FONT, size=10)
KEY_FONT = Font(name=FONT, bold=True, size=10, color="1F3864")
HEAD_FILL = PatternFill("solid", fgColor="1F3864")
KEY_FILL = PatternFill("solid", fgColor="E8ECF6")
_thin = Side(style="thin", color="C6CEDD")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
WRAP = Alignment(wrap_text=True, vertical="top")
HEAD_AL = Alignment(wrap_text=True, vertical="center", horizontal="center")


def add(wb, name, title, headers, rows, widths=None):
    ws = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    for j, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=j, value=h)
        c.font, c.fill, c.alignment, c.border = HEAD_FONT, HEAD_FILL, HEAD_AL, BORDER
    for i, row in enumerate(rows, 4):
        for j, val in enumerate(row, 1):
            c = ws.cell(row=i, column=j, value=val)
            c.font, c.alignment, c.border = BODY_FONT, WRAP, BORDER
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:{get_column_letter(len(headers))}{3 + len(rows)}"
    for j in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(j)].width = (
            widths[j - 1] if widths and j - 1 < len(widths) else 22
        )
    ws.row_dimensions[1].height = 22
    return ws


def cover(wb, name, title, pairs):
    ws = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    for i, (k, v) in enumerate(pairs, 3):
        kc = ws.cell(row=i, column=1, value=k)
        kc.font, kc.fill, kc.alignment, kc.border = KEY_FONT, KEY_FILL, WRAP, BORDER
        vc = ws.cell(row=i, column=2, value=v)
        vc.font, vc.alignment, vc.border = BODY_FONT, WRAP, BORDER
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 92
    return ws


wb = Workbook()
wb.remove(wb.active)

# ─────────────────────────── 00_表紙 ───────────────────────────
cover(wb, "00_表紙", "AI人材採用マッチングシステム 基本設計書", [
    ("文書名", "AI人材採用マッチングシステム 基本設計書(Excel形式)"),
    ("対象システム", "AI人材採用マッチングシステム(求職者と企業をAIで結ぶ採用マッチングSaaS)"),
    ("対象リポジトリ", "ai-recruitment-system(npm workspaces モノレポ: @ars/shared, @ars/api, @ars/web)"),
    ("作成日", "2026-06-20"),
    ("作成者", "ケイスタンプ株式会社"),
    ("対象バージョン", "0.1.0"),
    ("Git branch", "feat/ai-recruitment-mvp"),
    ("Git commit", "1a46f42"),
    ("文書目的",
     "本システムの構成・画面・API・データベース・RLS・AIマッチング・多言語対応・テストを、"
     "要求範囲(FR-01〜FR-10)と突合して基本設計として体系化することを目的とする。本書はレビュー・"
     "設計レビュー及び課題管理に用いる。本書はケイスタンプ株式会社が作成した。"),
    ("対象範囲",
     "FR-01 認証・ユーザー管理 / FR-02 履歴書アップロード・解析 / FR-03 AIスキル分析 / "
     "FR-04 企業・求人管理 / FR-05 AIマッチング / FR-06 人材検索・候補者詳細 / FR-07 企業・求人検索 / "
     "FR-08 メッセージ・通知 / FR-09 多言語対応(ja/en/zh-CN/zh-TW) / FR-10 候補者比較・面接・採用フロー。"),
    ("前提・制約",
     "本書はケイスタンプ株式会社が、実装済みシステムと要求仕様(FR-01〜FR-10)を突合して作成した"
     "基本設計書である。要求仕様書の細目と個別に突合できていない事項は『要確認』として要件担当と"
     "確認する。記載はいずれも実装事実に基づき、未実装の挙動は記載しない。"
     ""
     ""),
])

# ─────────────────────────── 01_改訂履歴 ───────────────────────────
add(wb, "01_改訂履歴", "01_改訂履歴",
    ["版数", "日付", "変更区分", "変更内容", "作成者", "レビュー状態"],
    [["1.0", "2026-06-20", "新規作成",
      "初版。実装内容と要求範囲(FR-01〜FR-10)の突合に基づき作成した。",
      "ケイスタンプ株式会社", "未レビュー(要レビュー)"]],
    [8, 12, 12, 70, 16, 16])

# ─────────────────────────── 02_凡例_文書構成 ───────────────────────────
add(wb, "02_凡例_文書構成", "02_凡例・文書構成",
    ["項目", "説明", "使用シート", "ステータス定義", "備考"],
    [
        ["差異分類:一致", "要求範囲と実装が整合している状態である。", "04,19,20", "一致", "—"],
        ["差異分類:要求のみ（未実装）", "要求範囲に含まれるが実装が確認できない状態である。", "19,20", "要求のみ（未実装）", "本調査では該当なし"],
        ["差異分類:実装のみ（要求書未記載）", "実装に存在するが要求列挙には明記されない機能である。", "19,20", "実装のみ（要求書未記載）", "ジョブキュー等の産業化要素"],
        ["差異分類:部分実装", "要求の一部のみ実装されている状態である。", "19,20", "部分実装", "—"],
        ["差異分類:仕様差異", "要求と実装で意味・挙動が異なる状態である。", "19,20", "仕様差異", "—"],
        ["差異分類:確認不能", "根拠資料の不足により判定できない状態である。", "19,20", "確認不能", "要件担当との確認待ち項目を含む"],
        ["実装状態", "各機能の実装の有無を示す。", "04,15", "実装済 / 一部 / 未実装", "—"],
        ["証跡の表記", "要求根拠・実装根拠・テスト根拠をパスで示す。", "全シート", "—",
         "実装根拠は apps/…、テスト根拠は *.test.ts / *.spec.ts、要求根拠は docx#FR で表す"],
        ["である調", "本書は敬体を用いず一貫して常体(である調)で記述する。", "全シート", "—", "—"],
        ["要求仕様の扱い", "要求仕様(FR-01〜FR-10)を要求範囲の基準とする。細目未突合は要確認とする。",
         "00,19,20,21", "—", "細目未突合は要確認"],
    ],
    [26, 46, 12, 24, 30])

# ─────────────────────────── 03_システム概要 ───────────────────────────
add(wb, "03_システム概要", "03_システム概要",
    ["区分", "設計内容", "要件根拠", "実装根拠", "補足"],
    [
        ["システム概要",
         "求職者の履歴書をAIで解析してスキルを抽出し、求人と候補者を双方向にマッチングするSaaSである。"
         "求職者・企業の二者が利用し、推薦・検索・メッセージ・採用フローを提供する。",
         "docx#概要 / FR-01〜FR-10", "README.md, docs/ARCHITECTURE.md", "—"],
        ["対象ユーザー", "求職者(job_seeker)と企業担当者(company_member)の二ロールである。未認証は公開閲覧のみ可能である。",
         "docx#対象ユーザー", "packages/shared(UserRole), apps/api/src/routes/auth.ts", "anonは公開求人・企業のみ閲覧可"],
        ["ロール定義",
         "job_seeker(求職者)・company_member(企業担当者)・anon(未認証)・service(内部処理)の4区分である。",
         "FR-01.3", "supabase/migrations/0001(profiles.role), 0009_rls.sql", "serviceはRLSを迂回する信頼経路"],
        ["全体構成",
         "React 18 SPA(Vite)+ Fastify 5 API + PostgreSQL(pgvector)から成る。ARS_RUNTIMEにより"
         "local(PGlite+モック)とsupabase(本番)の二系統で同一コードを動作させる。",
         "docx#アーキテクチャ", "apps/web, apps/api, supabase/migrations, apps/api/src/config.ts", "—"],
        ["フロントエンド境界",
         "react-router によるSPA、react-queryでサーバ状態を管理し、react-i18nextで多言語化する。"
         "デザインシステム(トークン+Radix)とAppShellで画面を構成する。",
         "FR-09", "apps/web/src/{App.tsx,design,components,i18n}", "—"],
        ["バックエンド境界",
         "Fastify。/api配下にルートを公開し、Zodで入出力を検証、リクエスト毎にRLSコンテキストを"
         "張る。helmet/レート制限/相関ID/OpenAPIを備える。",
         "FR-01〜FR-10", "apps/api/src/{app.ts,routes,services}", "—"],
        ["データベース境界",
         "PostgreSQL 26テーブル。全テーブルでRLSを有効化し、pgvectorで埋め込み近傍探索を行う。"
         "マイグレーションは前進専用(0001〜0012)である。",
         "FR-04,FR-05", "supabase/migrations/0001〜0012", "ivfflat(cosine, 384次元)"],
        ["AI境界",
         "LLMスキル分析と埋め込み生成をアダプタで抽象化し、mock/anthropic/openaiを切替える。"
         "localではdeterministicモックで外部依存なしに動作する。",
         "FR-03,FR-05", "apps/api/src/adapters/ai/*", "—"],
        ["ストレージ境界",
         "履歴書ファイルは本番ではSupabase Storageの非公開バケットresumesに、localではファイルシステムに"
         "保存する。所有者パス分離+API仲介でアクセス制御する。",
         "FR-02.3", "supabase/migrations/0010_storage.sql, apps/api/src/services/resume.ts", "—"],
    ],
    [16, 56, 22, 40, 26])

# ─────────────────────────── 04_機能一覧 ───────────────────────────
add(wb, "04_機能一覧", "04_機能一覧(FR-01〜FR-10)",
    ["機能ID", "機能名", "概要", "対象ロール", "主要画面", "主要API", "主要DB", "実装状態",
     "要件根拠", "実装根拠", "テスト根拠", "備考"],
    [
        ["FR-01", "認証・ユーザー管理",
         "登録・ログイン・メール確認・アカウント更新を行う。ロールはjob_seeker/company_memberである。",
         "全ロール", "/login,/register,/verify,/settings", "POST /api/auth/*, PATCH /api/auth/account",
         "profiles, candidates", "実装済", "docx#FR-01", "apps/api/src/routes/auth.ts",
         "apps/api/tests/auth.integration.test.ts", "JWT(local)/GoTrue(supabase)"],
        ["FR-02", "履歴書アップロード・解析",
         "PDF/DOCXを検証・ウイルススキャンし、解析ジョブを起動してテキスト抽出と解析を行う。",
         "job_seeker", "/me", "POST /api/candidates/me/resume, GET …/parse-jobs/:id",
         "resume_files, parse_jobs, job_queue", "実装済", "docx#FR-02",
         "apps/api/src/services/resume.ts", "apps/api/tests/resume.integration.test.ts", "10MB上限/マジックバイト検査"],
        ["FR-03", "AIスキル分析",
         "解析結果からスキル・熟練度・経験年数・キャリア提案・推奨学習を生成し、レーダーで可視化する。",
         "job_seeker", "/me", "GET/POST /api/candidates/me/analysis",
         "skill_analyses, candidate_skills", "実装済", "docx#FR-03",
         "apps/api/src/services/analysis.ts", "apps/api/tests/matching.integration.test.ts", "Zod検証+再試行"],
        ["FR-04", "企業・求人管理",
         "企業の作成・更新、求人のCRUDと公開範囲(draft/open, public/private)管理を行う。",
         "company_member", "/console,/console/companies/:id,/console/jobs/:id",
         "POST/PATCH /api/companies, /api/companies/:id/jobs, PATCH/DELETE /api/jobs/:id",
         "companies, company_members, jobs, job_skills", "実装済", "docx#FR-04",
         "apps/api/src/routes/{company,job}.ts", "apps/api/tests/company-job.integration.test.ts", "—"],
        ["FR-05", "AIマッチング",
         "埋め込み近傍探索とルールベース加重(match-v1)でスコア0〜100・理由・一致/不足スキルを算出する。",
         "job_seeker/company_member", "/recommendations, /jobs/:id(候補者)",
         "GET /api/candidates/me/recommendations, GET /api/jobs/:id/candidates",
         "match_results, *_embeddings, algorithm_versions", "実装済", "docx#FR-05",
         "apps/api/src/services/matching.ts, packages/shared/src/scoring.ts",
         "apps/api/tests/matching.integration.test.ts, packages/shared/src/scoring.test.ts", "決定論的・版管理"],
        ["FR-06", "人材検索・候補者詳細",
         "企業が候補者を検索・絞り込み・比較し、詳細を閲覧する。連絡先・履歴書は関係成立時のみ開示する。",
         "company_member", "/talent,/talent/:id",
         "GET /api/talent, GET /api/talent/:id, POST /api/compare",
         "candidates, candidate_skills, shortlists, match_results", "実装済", "docx#FR-06",
         "apps/api/src/services/talent.ts", "apps/api/tests/talent.integration.test.ts", "candidate_readableで可視性制御"],
        ["FR-07", "企業・求人検索",
         "公開求人・公開企業の検索・絞り込み・ページングを未認証でも行える。",
         "全ロール", "/jobs,/companies,/companies/:id",
         "GET /api/jobs, GET /api/companies, GET /api/companies/:id/jobs",
         "jobs, companies, job_skills", "実装済", "docx#FR-07",
         "apps/api/src/routes/{job,company}.ts", "apps/api/tests/company-job.integration.test.ts", "anonはopen+publicのみ"],
        ["FR-08", "メッセージ・通知",
         "求職者と企業のスレッド型メッセージ、未読管理、各種イベント通知(応募・面接等)を提供する。",
         "全認証ロール", "/messages,/notifications",
         "POST/GET /api/conversations*, GET/POST /api/notifications*",
         "conversations, conversation_members, messages, notifications", "実装済", "docx#FR-08",
         "apps/api/src/routes/messaging.ts", "apps/api/tests/messaging.integration.test.ts", "client_tokenで重複排除"],
        ["FR-09", "多言語対応",
         "ja/en/zh-CN/zh-TWの4ロケールに対応し、UIとAPIエラーを翻訳する。辞書はキー整合をテストで担保する。",
         "全ロール", "全画面(言語切替)",
         "Accept-Language/messageKey連携", "—(クライアント辞書)", "実装済", "docx#FR-09",
         "apps/web/src/i18n/*", "apps/web/src/i18n/i18n.parity.test.ts", "zh-*は遅延ロード"],
        ["FR-10", "候補者比較・面接・採用フロー",
         "応募の段階遷移(applied〜hired)、ショートリスト、候補者比較、面接提案・応答を管理する。",
         "job_seeker/company_member", "/applications,/talent,/console/jobs/:id",
         "POST /api/applications, PATCH …/stage, POST …/interviews, POST /api/shortlists, /api/compare",
         "applications, application_stage_history, interviews, shortlists, candidate_comparisons",
         "実装済", "docx#FR-10", "apps/api/src/routes/recruitment.ts",
         "apps/api/tests/recruitment.integration.test.ts", "段階遷移はRLS+サービスで権限制御"],
    ],
    [8, 18, 40, 16, 22, 30, 24, 10, 14, 28, 30, 18])

# ─────────────────────────── 05_業務フロー ───────────────────────────
add(wb, "05_業務フロー", "05_業務フロー",
    ["フローID", "フロー名", "アクター", "開始条件", "ステップ", "ユーザー操作/システム処理",
     "主な画面", "主なAPI", "成功条件", "例外・分岐", "証跡"],
    [
        ["F-01", "求職者の入職フロー", "求職者", "未登録である", "1", "登録・メール確認を行う", "/register,/verify",
         "POST /api/auth/register, /verify-email", "セッションを取得する", "重複メールはCONFLICT", "auth.ts"],
        ["F-01", "求職者の入職フロー", "求職者", "認証済である", "2", "履歴書(PDF/DOCX)をアップロードする", "/me",
         "POST /api/candidates/me/resume", "解析ジョブが起動する", "サイズ超過/形式不正/感染で拒否", "resume.ts"],
        ["F-01", "求職者の入職フロー", "システム", "解析ジョブが存在する", "3", "テキスト抽出とAIスキル分析を実行する", "—",
         "(job)resume_parse", "skill_analysesを生成する", "失敗時はリトライ後dead", "analysis.ts, jobs/queue.ts"],
        ["F-01", "求職者の入職フロー", "求職者", "分析が完了している", "4", "推薦求人とスコアを閲覧する", "/recommendations",
         "GET /api/candidates/me/recommendations", "推薦一覧を表示する", "未分析時は空状態", "match.ts"],
        ["F-01", "求職者の入職フロー", "求職者", "求人を選択した", "5", "求人へ応募する", "/jobs/:id,/applications",
         "POST /api/applications", "応募が登録される", "二重応募はCONFLICT", "recruitment.ts"],
        ["F-01", "求職者の入職フロー", "求職者", "面接提案を受けた", "6", "面接を確認/辞退する", "/applications",
         "POST /api/interviews/:id/respond", "面接状態が更新される", "—", "recruitment.ts"],
        ["F-02", "企業の採用フロー", "企業担当者", "認証済である", "1", "企業を作成する", "/console",
         "POST /api/companies", "企業とメンバーが作られる", "company_member以外は不可", "company.ts"],
        ["F-02", "企業の採用フロー", "企業担当者", "企業が存在する", "2", "求人を作成・公開する", "/console/companies/:id",
         "POST /api/companies/:id/jobs", "求人が公開される", "範囲外編集はFORBIDDEN", "job.ts"],
        ["F-02", "企業の採用フロー", "企業担当者", "求人が公開済である", "3", "人材を検索・比較する", "/talent",
         "GET /api/talent, POST /api/compare", "候補者一覧/比較を得る", "可視性外は除外", "talent.ts"],
        ["F-02", "企業の採用フロー", "企業担当者", "候補者を選定した", "4", "ショートリスト追加・応募者選考を行う", "/console/jobs/:id",
         "POST /api/shortlists, PATCH …/stage", "段階が更新される", "段階遷移権限を検証", "recruitment.ts"],
        ["F-02", "企業の採用フロー", "企業担当者", "選考が進んだ", "5", "面接を提案し採用判断する", "/console/jobs/:id",
         "POST …/interviews, PATCH …/stage(hired)", "採用に至る", "—", "recruitment.ts"],
    ],
    [8, 18, 14, 16, 6, 30, 18, 30, 22, 22, 18])

# ─────────────────────────── 06_画面一覧 ───────────────────────────
SCR = [
    ["SC-01", "ホーム", "/", "公開", "ランディング。ロール別の導線を提示する。", "登録/ログイン/検索への導線",
     "—", "home.*, nav.*", "なし", "—", "apps/web/src/pages/Home.tsx", "—"],
    ["SC-02", "ログイン", "/login", "公開", "メールとパスワードで認証する。", "認証, エラー表示",
     "POST /api/auth/login", "auth.loginTitle ほか", "なし", "Loading/Error",
     "apps/web/src/pages/Login.tsx", "apps/web/src/pages/Login.test.tsx"],
    ["SC-03", "新規登録", "/register", "公開", "ロールを選んで登録する。", "登録, バリデーション",
     "POST /api/auth/register", "auth.register*", "なし", "Error/項目エラー", "apps/web/src/pages/Register.tsx", "—"],
    ["SC-04", "メール確認", "/verify", "公開", "確認トークンを入力する。", "確認",
     "POST /api/auth/verify-email", "auth.verifyEmail*", "なし", "Loading/Error", "apps/web/src/pages/VerifyEmail.tsx", "—"],
    ["SC-05", "求人一覧", "/jobs", "公開", "公開求人を検索・絞り込みする。", "検索, フィルタ, ページング",
     "GET /api/jobs", "job.*", "なし", "Loading/Empty/Error",
     "apps/web/src/pages/JobsBrowse.tsx", "apps/web/src/pages/JobsBrowse.test.tsx"],
    ["SC-06", "求人詳細", "/jobs/:id", "公開", "求人詳細を表示し求職者は応募する。", "応募",
     "GET /api/jobs/:id, POST /api/applications", "job.*, application.apply", "応募はjob_seekerのみ", "Loading/Error",
     "apps/web/src/pages/JobDetail.tsx", "—"],
    ["SC-07", "企業一覧", "/companies", "公開", "公開企業を検索する。", "検索, フィルタ",
     "GET /api/companies", "company.*", "なし", "Loading/Empty/Error", "apps/web/src/pages/CompaniesBrowse.tsx", "—"],
    ["SC-08", "企業詳細", "/companies/:id", "公開", "企業情報と求人を表示する。", "求人一覧",
     "GET /api/companies/:id, …/jobs", "company.*", "なし", "Loading/Error", "apps/web/src/pages/CompanyDetail.tsx", "—"],
    ["SC-09", "アカウント設定", "/settings", "認証", "表示名・言語・パスワードを更新する。", "更新, 言語切替, 離脱警告",
     "PATCH /api/auth/account", "auth.accountSettings ほか", "ProtectedRoute", "Loading/Done/Error",
     "apps/web/src/pages/AccountSettings.tsx", "apps/web/src/hooks/useUnsavedChangesWarning.test.tsx"],
    ["SC-10", "メッセージ", "/messages", "認証", "会話一覧とスレッドを表示し送信する。", "送信, 未読管理",
     "GET/POST /api/conversations*", "messaging.*", "ProtectedRoute", "Loading/Empty", "apps/web/src/pages/Messages.tsx", "—"],
    ["SC-11", "通知", "/notifications", "認証", "通知一覧と既読化を行う。", "既読, 一括既読",
     "GET/POST /api/notifications*", "notification.*", "ProtectedRoute", "Loading/Empty", "apps/web/src/pages/Notifications.tsx", "—"],
    ["SC-12", "履歴書・スキル", "/me", "求職者", "履歴書アップロードと分析・レーダーを表示する。", "アップロード, 再試行, 再生成",
     "GET /api/candidates/me, …/analysis, POST …/resume", "resume.*, analysis.*", "ProtectedRoute(job_seeker)",
     "Loading/Empty/Error/処理中", "apps/web/src/pages/SeekerProfile.tsx",
     "apps/web/src/components/ResumeDropzone.test.tsx, RadarChart.test.tsx"],
    ["SC-13", "おすすめ", "/recommendations", "求職者", "推薦求人とスコアを表示する。", "スコア・理由表示",
     "GET /api/candidates/me/recommendations", "match.*", "ProtectedRoute(job_seeker)", "Loading/Empty/Error",
     "apps/web/src/pages/Recommendations.tsx", "—"],
    ["SC-14", "応募一覧", "/applications", "求職者", "応募状況と段階を表示し辞退する。", "段階表示, 辞退",
     "GET /api/applications, PATCH …/stage", "application.*", "ProtectedRoute(job_seeker)", "Loading/Empty/Error",
     "apps/web/src/pages/MyApplications.tsx", "—"],
    ["SC-15", "企業コンソール", "/console", "企業", "所有企業の一覧と新規作成を行う。", "企業作成",
     "GET /api/companies/mine, POST /api/companies", "company.*", "ProtectedRoute(company_member)", "Loading/Empty/Error",
     "apps/web/src/pages/CompanyConsole.tsx", "—"],
    ["SC-16", "企業管理", "/console/companies/:id", "企業", "企業編集と求人管理を行う。", "企業更新, 求人CRUD",
     "PATCH /api/companies/:id, …/manage/jobs", "company.*, job.*", "ProtectedRoute(company_member)", "Loading/Error",
     "apps/web/src/pages/CompanyManage.tsx", "—"],
    ["SC-17", "求人管理", "/console/jobs/:id", "企業", "求人編集と応募者選考・面接を行う。", "求人更新, 段階遷移, 面接提案",
     "PATCH /api/jobs/:id, …/applications, …/interviews", "job.*, application.*", "ProtectedRoute(company_member)",
     "Loading/Error", "apps/web/src/pages/JobManage.tsx", "—"],
    ["SC-18", "人材検索", "/talent", "企業", "候補者を検索・絞り込み・比較する。", "検索, 比較, ショートリスト",
     "GET /api/talent, POST /api/compare, /shortlists", "talent.*, shortlist.*", "ProtectedRoute(company_member)",
     "Loading/Empty/Error", "apps/web/src/pages/TalentSearch.tsx", "—"],
    ["SC-19", "候補者詳細", "/talent/:id", "企業", "候補者詳細を表示しショートリストに追加する。", "詳細表示, ショートリスト",
     "GET /api/talent/:id, POST /api/shortlists", "candidate.*", "ProtectedRoute(company_member)", "Loading/Error",
     "apps/web/src/pages/CandidateDetail.tsx", "—"],
    ["SC-20", "ショートリスト", "/shortlist", "企業", "保存候補者の一覧と削除を行う。", "一覧, 削除",
     "GET /api/shortlists, DELETE …/:id", "shortlist.*", "ProtectedRoute(company_member)", "Loading/Empty/Error",
     "apps/web/src/pages/Shortlist.tsx", "—"],
    ["SC-21", "アクセス拒否(403)", "/403", "全", "権限不足を明示する状態画面である。", "ホーム導線",
     "—", "errorPage.forbidden*", "なし", "—", "apps/web/src/pages/Forbidden.tsx", "apps/web/e2e/a11y.spec.ts"],
    ["SC-22", "サーバエラー(500)", "/500", "全", "描画失敗時にErrorBoundaryが表示する。", "再読込導線",
     "—", "errorPage.server*", "ErrorBoundary", "—", "apps/web/src/pages/ServerError.tsx", "—"],
    ["SC-23", "未検出(404)", "*", "全", "未定義経路のキャッチオールである。", "ホーム導線",
     "—", "errorPage.notFound*", "なし", "—", "apps/web/src/pages/NotFound.tsx", "apps/web/e2e/a11y.spec.ts"],
]
add(wb, "06_画面一覧", "06_画面一覧",
    ["画面ID", "画面名", "URL/Route", "対象ロール", "概要", "主要機能", "主要API",
     "多言語キー/i18n根拠", "権限制御", "状態(Loading/Empty/Error)", "実装根拠", "テスト根拠"],
    SCR, [8, 16, 20, 10, 30, 22, 30, 18, 20, 18, 30, 26])

# ─────────────────────────── 07_画面項目定義 ───────────────────────────
add(wb, "07_画面項目定義", "07_画面項目定義(主要画面)",
    ["画面ID", "項目ID", "項目名", "表示/入力", "データ型", "必須", "入力制約", "バリデーション",
     "表示条件", "保存先/API項目", "エラーメッセージ", "実装根拠"],
    [
        ["SC-02", "I-01", "メールアドレス", "入力", "string", "必須", "メール形式", "Zod email", "常時",
         "POST /api/auth/login(email)", "auth.emailField.invalid", "apps/web/src/pages/Login.tsx"],
        ["SC-02", "I-02", "パスワード", "入力", "string", "必須", "8文字以上・英数字", "Zod password",
         "常時", "POST /api/auth/login(password)", "auth.password.*", "apps/web/src/pages/Login.tsx"],
        ["SC-03", "I-03", "ロール", "入力", "enum", "必須", "job_seeker/company_member", "Zod enum", "常時",
         "POST /api/auth/register(role)", "—", "apps/web/src/pages/Register.tsx"],
        ["SC-03", "I-04", "表示名", "入力", "string", "必須", "1文字以上", "Zod min(1)", "常時",
         "register(displayName)", "common.requiredField", "apps/web/src/pages/Register.tsx"],
        ["SC-12", "I-05", "履歴書ファイル", "入力", "file(PDF/DOCX)", "必須", "10MB以下・拡張子/MIME/マジックバイト整合",
         "validateUpload + サーバ検査", "常時", "POST /api/candidates/me/resume(file)",
         "resume.upload.{tooLarge,badExtension,badMime,badContent,infected}", "apps/web/src/components/ResumeDropzone.tsx"],
        ["SC-12", "I-06", "経験年数", "入力", "number", "任意", "0〜60", "min/max", "プロフィール編集時",
         "PATCH /api/candidates/me(yearsExperience)", "—", "apps/web/src/pages/SeekerProfile.tsx"],
        ["SC-17", "I-07", "求人タイトル", "入力", "string", "必須", "1文字以上", "CreateJobSchema", "求人作成/編集",
         "POST/PATCH /api/jobs(title)", "common.requiredField", "apps/web/src/pages/JobManage.tsx"],
        ["SC-17", "I-08", "必須スキル", "入力", "string[]", "任意", "カンマ区切り", "配列正規化", "求人作成/編集",
         "jobs.job_skills(required)", "—", "apps/web/src/pages/JobManage.tsx"],
        ["SC-17", "I-09", "給与下限/上限", "入力", "number", "任意", "上限≧下限", "Zod refine + DB Check",
         "求人作成/編集", "jobs.salary_min/max", "job.salary.rangeInvalid", "apps/web/src/pages/JobManage.tsx"],
        ["SC-18", "I-10", "検索キーワード", "入力", "string", "任意", "—", "—", "常時",
         "GET /api/talent(q)", "—", "apps/web/src/pages/TalentSearch.tsx"],
        ["SC-18", "I-11", "最小経験年数", "入力", "number", "任意", "0以上", "Zod coerce", "常時",
         "GET /api/talent(minYears)", "—", "apps/web/src/pages/TalentSearch.tsx"],
        ["SC-14", "I-12", "応募段階", "表示", "enum", "—", "applied〜hired/rejected/withdrawn", "—", "常時",
         "applications.stage", "—", "apps/web/src/pages/MyApplications.tsx"],
    ],
    [8, 7, 16, 10, 14, 8, 24, 20, 14, 28, 26, 30])

# ─────────────────────────── 08_API一覧 ───────────────────────────
API = [
    ["A-01", "POST", "/api/auth/register", "ユーザー登録を行う", "不要", "—", "RegisterSchema", "Session",
     "profiles, candidates(insert)", "なし(登録経路)", "VALIDATION, CONFLICT", "auth.ts:13", "auth.integration.test.ts"],
    ["A-02", "POST", "/api/auth/login", "認証してトークンを発行する", "不要", "—", "LoginSchema", "Session",
     "profiles(select)", "なし", "UNAUTHORIZED, VALIDATION", "auth.ts:18", "auth.integration.test.ts"],
    ["A-03", "POST", "/api/auth/verify-email", "メール確認を行う", "不要", "—", "VerifyEmailSchema", "Session",
     "auth(identities)", "—", "NOT_FOUND, VALIDATION", "auth.ts:22", "auth.integration.test.ts"],
    ["A-04", "GET", "/api/auth/me", "認証中の自プロフィールを返す", "要", "全", "—", "AuthUser",
     "profiles(select)", "あり", "UNAUTHORIZED, NOT_FOUND", "auth.ts:29", "auth.integration.test.ts"],
    ["A-05", "PATCH", "/api/auth/account", "表示名・言語・パスワードを更新する", "要", "全", "UpdateAccountSchema", "AuthUser",
     "profiles(update)", "あり(id=auth.uid())", "UNAUTHORIZED, VALIDATION", "auth.ts:31", "auth.integration.test.ts"],
    ["A-06", "GET", "/api/candidates/me", "候補者プロフィールを返す", "要", "job_seeker", "—", "CandidateProfile",
     "candidates, candidate_skills(select)", "あり(owns)", "FORBIDDEN, NOT_FOUND", "candidate.ts:32", "resume.integration.test.ts"],
    ["A-07", "PATCH", "/api/candidates/me", "候補者プロフィールを更新する", "要", "job_seeker", "UpdateCandidateProfileSchema",
     "CandidateProfile", "candidates(update), embeddings(再生成)", "あり(owns)", "VALIDATION", "candidate.ts:34", "resume.integration.test.ts"],
    ["A-08", "POST", "/api/candidates/me/resume", "履歴書をアップロードし解析を起動する", "要", "job_seeker", "multipart(file)",
     "UploadResult", "resume_files, parse_jobs(insert), job_queue", "あり(owns)",
     "PAYLOAD_TOO_LARGE, UNSUPPORTED_MEDIA_TYPE, VIRUS_DETECTED, VALIDATION", "candidate.ts:42", "resume.integration.test.ts"],
    ["A-09", "GET", "/api/candidates/me/resumes", "履歴書一覧を返す", "要", "job_seeker", "—", "ResumeFile[]",
     "resume_files(select)", "あり(owns)", "FORBIDDEN", "candidate.ts:55", "resume.integration.test.ts"],
    ["A-10", "GET", "/api/candidates/me/parse-jobs/:id", "解析ジョブ状態を返す", "要", "job_seeker", "—", "ParseJob",
     "parse_jobs(select)", "あり(owns)", "NOT_FOUND", "candidate.ts:59", "resume.integration.test.ts"],
    ["A-11", "POST", "/api/candidates/me/parse-jobs/:id/retry", "失敗解析を再実行する", "要", "job_seeker", "—", "ParseJob",
     "parse_jobs, job_queue", "あり(owns)", "NOT_FOUND, CONFLICT", "candidate.ts:63", "resume.integration.test.ts"],
    ["A-12", "GET", "/api/candidates/me/analysis", "最新スキル分析を返す", "要", "job_seeker", "—", "SkillAnalysis",
     "skill_analyses(select)", "あり(owns)", "NOT_FOUND(analysis.none)", "candidate.ts:67", "resume.integration.test.ts"],
    ["A-13", "POST", "/api/candidates/me/analysis", "スキル分析を生成/再生成する", "要", "job_seeker", "GenerateAnalysisSchema",
     "SkillAnalysis", "skill_analyses, job_queue", "あり(owns)", "UPSTREAM_AI_ERROR, VALIDATION", "candidate.ts:79", "resume.integration.test.ts"],
    ["A-14", "GET", "/api/resumes/:id/download", "履歴書ファイルをダウンロードする", "要", "job_seeker/company_member", "—", "binary",
     "resume_files, applications(認可)", "あり(API仲介)", "FORBIDDEN, NOT_FOUND", "candidate.ts:88", "security-files.integration.test.ts"],
    ["A-15", "POST", "/api/companies", "企業を作成する", "要", "company_member", "CreateCompanySchema", "Company",
     "companies, company_members(insert)", "あり(role)", "VALIDATION, CONFLICT", "company.ts:22", "company-job.integration.test.ts"],
    ["A-16", "GET", "/api/companies", "公開企業を検索する", "不要", "全", "CompanySearchQuerySchema", "Paginated<Company>",
     "companies, jobs(select)", "あり(公開)", "VALIDATION", "company.ts:32", "company-job.integration.test.ts"],
    ["A-17", "GET", "/api/companies/mine", "所属企業一覧を返す", "要", "company_member", "—", "Company[]",
     "companies, company_members", "あり(member)", "FORBIDDEN", "company.ts:36", "company-job.integration.test.ts"],
    ["A-18", "GET", "/api/companies/:id", "企業詳細を返す", "不要", "全", "—", "Company", "companies(select)",
     "あり(公開)", "NOT_FOUND", "company.ts:40", "company-job.integration.test.ts"],
    ["A-19", "PATCH", "/api/companies/:id", "企業を更新する", "要", "company_member", "UpdateCompanySchema", "Company",
     "companies(update)", "あり(member)", "FORBIDDEN, NOT_FOUND, VALIDATION", "company.ts:44", "company-job.integration.test.ts"],
    ["A-20", "GET", "/api/companies/:id/members", "企業メンバーを返す", "要", "company_member", "—", "CompanyMember[]",
     "company_members, profiles", "あり(member)", "FORBIDDEN, NOT_FOUND", "company.ts:53", "company-job.integration.test.ts"],
    ["A-21", "GET", "/api/companies/:id/jobs", "企業の公開求人を返す", "不要", "全", "JobSearchQuerySchema", "Paginated<Job>",
     "jobs, job_skills", "あり(公開)", "VALIDATION", "company.ts:58", "company-job.integration.test.ts"],
    ["A-22", "GET", "/api/jobs", "公開求人を検索する", "不要", "全", "JobSearchQuerySchema", "Paginated<Job>",
     "jobs, job_skills, companies", "あり(open+public)", "VALIDATION", "job.ts:17", "company-job.integration.test.ts"],
    ["A-23", "GET", "/api/jobs/:id", "求人詳細を返す", "不要", "全", "—", "Job", "jobs, job_skills(select)",
     "あり(open+public/member)", "NOT_FOUND", "job.ts:21", "company-job.integration.test.ts"],
    ["A-24", "GET", "/api/companies/:companyId/manage/jobs", "管理用に全求人を返す", "要", "company_member", "—", "Job[]",
     "jobs(select)", "あり(member)", "FORBIDDEN, NOT_FOUND", "job.ts:26", "company-job.integration.test.ts"],
    ["A-25", "POST", "/api/companies/:companyId/jobs", "求人を作成する", "要", "company_member", "CreateJobSchema", "Job",
     "jobs, job_skills, job_embeddings", "あり(member)", "VALIDATION, NOT_FOUND", "job.ts:30", "company-job.integration.test.ts"],
    ["A-26", "PATCH", "/api/jobs/:id", "求人を更新する", "要", "company_member", "UpdateJobSchema", "Job",
     "jobs, job_skills, job_embeddings", "あり(member)", "FORBIDDEN, NOT_FOUND, VALIDATION", "job.ts:41", "company-job.integration.test.ts"],
    ["A-27", "DELETE", "/api/jobs/:id", "求人を削除する", "要", "company_member", "—", "204", "jobs(delete)",
     "あり(member)", "FORBIDDEN, NOT_FOUND", "job.ts:50", "company-job.integration.test.ts"],
    ["A-28", "GET", "/api/candidates/me/recommendations", "推薦求人を返す", "要", "job_seeker", "—", "MatchResult[]",
     "match_results, jobs", "あり(owns)", "FORBIDDEN", "match.ts:8", "matching.integration.test.ts"],
    ["A-29", "GET", "/api/jobs/:id/candidates", "求人の候補者ランキングを返す", "要", "company_member", "—", "MatchResult[]",
     "match_results, candidates", "あり(member, readable)", "FORBIDDEN, NOT_FOUND", "match.ts:13", "matching.integration.test.ts"],
    ["A-30", "GET", "/api/talent", "候補者を検索する", "要", "company_member", "TalentSearchQuerySchema", "Paginated<TalentSummary>",
     "candidates, candidate_skills", "あり(readable)", "VALIDATION", "talent.ts:10", "talent.integration.test.ts"],
    ["A-31", "GET", "/api/talent/:id", "候補者詳細を返す", "要", "company_member", "—", "CandidateDetail",
     "candidates, resume_files(条件付)", "あり(readable)", "NOT_FOUND", "talent.ts:19", "talent.integration.test.ts"],
    ["A-32", "POST", "/api/conversations", "会話を作成する", "要", "全", "CreateConversationSchema", "Conversation",
     "conversations, members, messages", "あり", "VALIDATION, NOT_FOUND", "messaging.ts:32", "messaging.integration.test.ts"],
    ["A-33", "POST", "/api/conversations/with-company", "企業との会話を開始/再利用する", "要", "job_seeker", "StartWithCompanySchema",
     "Conversation", "conversations, members", "あり", "VALIDATION, NOT_FOUND", "messaging.ts:42", "messaging.integration.test.ts"],
    ["A-34", "GET", "/api/conversations", "会話一覧を返す", "要", "全", "—", "Conversation[]",
     "conversations, members", "あり(member)", "UNAUTHORIZED", "messaging.ts:52", "messaging.integration.test.ts"],
    ["A-35", "GET", "/api/conversations/:id/messages", "メッセージを返す", "要", "全", "—", "Message[]",
     "messages", "あり(member)", "FORBIDDEN, NOT_FOUND", "messaging.ts:58", "messaging.integration.test.ts"],
    ["A-36", "POST", "/api/conversations/:id/messages", "メッセージを送信する", "要", "全", "SendMessageSchema", "Message",
     "messages, notifications", "あり(member)", "VALIDATION, FORBIDDEN, RATE_LIMITED", "messaging.ts:62", "messaging.integration.test.ts"],
    ["A-37", "GET", "/api/notifications", "通知一覧を返す", "要", "全", "—", "Notification[]", "notifications(select)",
     "あり(self)", "UNAUTHORIZED", "messaging.ts:95", "messaging.integration.test.ts"],
    ["A-38", "POST", "/api/notifications/:id/read, /read-all", "通知を既読化する", "要", "全", "—", "204",
     "notifications(update)", "あり(self)", "FORBIDDEN, NOT_FOUND", "messaging.ts:97", "messaging.integration.test.ts"],
    ["A-39", "POST", "/api/applications", "求人へ応募する", "要", "job_seeker", "CreateApplicationSchema", "Application",
     "applications(insert), notifications", "あり(owns)", "VALIDATION, NOT_FOUND, CONFLICT, RATE_LIMITED",
     "recruitment.ts:36", "recruitment.integration.test.ts"],
    ["A-40", "GET", "/api/applications, /jobs/:id/applications", "応募一覧を返す", "要", "job_seeker/company_member", "—",
     "Application[]", "applications, jobs", "あり(readable)", "FORBIDDEN", "recruitment.ts:44", "recruitment.integration.test.ts"],
    ["A-41", "PATCH", "/api/applications/:id/stage", "応募段階を更新する", "要", "job_seeker/company_member", "UpdateApplicationStageSchema",
     "Application", "applications, stage_history", "あり(readable+権限)", "VALIDATION, FORBIDDEN, CONFLICT",
     "recruitment.ts:50", "recruitment.integration.test.ts"],
    ["A-42", "POST", "/api/shortlists, /compare", "ショートリスト追加/候補者比較を行う", "要", "company_member",
     "AddShortlistSchema/CompareRequestSchema", "ShortlistEntry/ComparisonResult", "shortlists, candidate_comparisons",
     "あり(member)", "VALIDATION, NOT_FOUND, CONFLICT", "recruitment.ts:60,73", "recruitment.integration.test.ts"],
    ["A-43", "POST", "/api/applications/:id/interviews, /interviews/:id/respond", "面接提案/応答を行う", "要",
     "company_member/job_seeker", "ProposeInterviewSchema/RespondInterviewSchema", "Interview", "interviews, notifications",
     "あり(readable)", "VALIDATION, FORBIDDEN, RATE_LIMITED", "recruitment.ts:82,98", "recruitment.integration.test.ts"],
    ["A-44", "GET", "/health, /ready, /metrics, /openapi.json", "稼働/受入/計測/契約を公開する", "不要", "全", "—",
     "health/text/json", "—(/readyはDB往復)", "503(ready未準備時)", "app.ts", "health.integration.test.ts, openapi.integration.test.ts"],
]
add(wb, "08_API一覧", "08_API一覧",
    ["API ID", "Method", "Path", "概要", "認証要否", "許可ロール", "Request Schema", "Response Schema",
     "主なDB操作", "RLS影響", "エラーコード", "実装根拠", "テスト根拠"],
    API, [7, 8, 34, 26, 8, 18, 24, 22, 26, 18, 26, 16, 26])

# ─────────────────────────── 09_API詳細 ───────────────────────────
add(wb, "09_API詳細", "09_API詳細(代表API)",
    ["API ID", "区分", "項目名", "型", "必須", "制約", "説明", "正常時", "異常時", "備考"],
    [
        ["A-02", "Request", "email", "string", "必須", "メール形式", "ログインメール", "—", "VALIDATION(422)", "—"],
        ["A-02", "Request", "password", "string", "必須", "—", "パスワード", "—", "UNAUTHORIZED(401)", "ログ非出力"],
        ["A-02", "Response", "accessToken", "string", "—", "JWT", "アクセストークン", "200で返す", "—", "Bearerで利用"],
        ["A-02", "Response", "user", "AuthUser", "—", "—", "認証ユーザー情報", "200で返す", "—", "—"],
        ["A-08", "Request", "file", "multipart", "必須", "PDF/DOCX・10MB以下", "履歴書ファイル", "201で受理", "413/415/422", "マジックバイト検査"],
        ["A-08", "副作用", "parse_jobs", "row", "—", "—", "解析ジョブを起動する", "job_queueへ投入", "—", "localは即時drain"],
        ["A-08", "Response", "parseJob.status", "enum", "—", "pending..", "解析状態", "pendingを返す", "—", "クライアントがポーリング"],
        ["A-08", "異常", "ウイルス検知", "—", "—", "—", "感染時は保存しない", "—", "VIRUS_DETECTED(422)", "resume.upload.infected"],
        ["A-39", "Request", "jobId", "uuid", "必須", "公開求人", "応募対象求人", "201で登録", "NOT_FOUND/VALIDATION", "—"],
        ["A-39", "Request", "coverNote", "string", "任意", "—", "カバーレター", "—", "—", "PII扱い"],
        ["A-39", "異常", "二重応募", "—", "—", "unique(job,candidate)", "重複応募を拒否する", "—", "CONFLICT(409)", "application.duplicate"],
        ["A-28", "Response", "score", "integer", "—", "0〜100", "マッチスコア", "降順で返す", "—", "match_resultsから取得"],
        ["A-28", "Response", "reason/matched/missing", "—", "—", "—", "スコア説明と一致/不足スキル", "—", "—", "決定論的算定"],
    ],
    [7, 10, 18, 12, 8, 18, 28, 16, 22, 18])

# ─────────────────────────── 10_DBテーブル一覧 ───────────────────────────
TBL = [
    ["profiles", "認証ユーザーのメタ(表示名・ロール・ロケール)", "id", "—", "—", "全機能の基点", "有", "無", "0001", "PII(表示名)"],
    ["candidates", "求職者プロフィール", "id", "user_id→profiles(cascade)", "user_id", "FR-02/03/06", "有", "無", "0002", "PII(headline等)"],
    ["resume_files", "履歴書ファイルメタ", "id", "candidate_id→candidates(cascade)", "storage_path", "FR-02", "有", "無", "0002", "PII(抽出テキスト)"],
    ["parse_jobs", "履歴書解析ジョブ", "id", "candidate_id→candidates(cascade)", "—", "FR-02/03", "有", "無", "0002", "—"],
    ["skills", "スキル辞書", "id", "—", "name", "FR-03", "有", "無", "0002", "公開参照"],
    ["candidate_skills", "求職者スキル", "(candidate_id,skill_id)", "skill_id→skills(cascade)", "—", "FR-03", "有", "無", "0002", "—"],
    ["skill_analyses", "AIスキル分析結果", "id", "candidate_id→candidates(cascade)", "—", "FR-03", "有", "無", "0002", "PII(要約)"],
    ["companies", "企業プロフィール", "id", "—", "—", "FR-04/07", "有", "無", "0003", "—"],
    ["company_members", "企業メンバーシップ", "id", "company_id,user_id→profiles(cascade)", "(company_id,user_id)", "FR-04", "有", "無", "0003", "—"],
    ["jobs", "求人", "id", "company_id→companies(cascade)", "—", "FR-04/07", "有", "無", "0004", "Check(年数/給与)"],
    ["job_skills", "求人スキル要件", "(job_id,skill_name)", "job_id→jobs(cascade)", "—", "FR-04/05", "有", "無", "0004", "—"],
    ["algorithm_versions", "マッチング算定の版管理", "version", "—", "—", "FR-05", "有", "無", "0005", "—"],
    ["candidate_embeddings", "求職者埋め込み", "candidate_id", "candidate_id→candidates(cascade)", "—", "FR-05", "有", "vector(384)", "0005", "—"],
    ["job_embeddings", "求人埋め込み", "job_id", "job_id→jobs(cascade)", "—", "FR-05", "有", "vector(384)", "0005", "—"],
    ["match_results", "マッチング結果", "id", "job_id,candidate_id→candidates(cascade)", "(job_id,candidate_id)", "FR-05", "有", "無", "0006", "Check(0..100)"],
    ["conversations", "メッセージスレッド", "id", "job_id→jobs(set null);created_by→profiles", "—", "FR-08", "有", "無", "0007", "—"],
    ["conversation_members", "スレッドメンバー", "(conversation_id,user_id)", "→conversations/profiles(cascade)", "—", "FR-08", "有", "無", "0007", "—"],
    ["messages", "メッセージ本文", "id", "conversation_id,sender_user_id→profiles(cascade)", "(conv,sender,client_token)", "FR-08", "有", "無", "0007", "PII(本文)"],
    ["notifications", "通知", "id", "user_id→profiles(cascade)", "—", "FR-08", "有", "無", "0007", "PII(本文)"],
    ["applications", "応募", "id", "job_id,candidate_id→candidates(cascade)", "(job_id,candidate_id)", "FR-10", "有", "無", "0008", "PII(cover_note)"],
    ["application_stage_history", "応募段階履歴", "id", "application_id,changed_by→profiles(cascade)", "—", "FR-10", "有", "無", "0008", "—"],
    ["shortlists", "ショートリスト", "id", "company_id,candidate_id,created_by→profiles", "(company,candidate,job)", "FR-10", "有", "無", "0008", "—"],
    ["candidate_comparisons", "候補者比較", "id", "company_id,created_by→profiles(cascade)", "—", "FR-10", "有", "無", "0008", "—"],
    ["interviews", "面接", "id", "application_id,proposed_by→profiles(cascade)", "—", "FR-10", "有", "無", "0008", "PII(notes)"],
    ["job_queue", "耐久ジョブキュー", "id", "—", "idempotency_key", "内部処理(FR-02解析)", "有", "無", "0011", "service専用"],
]
add(wb, "10_DBテーブル一覧", "10_DBテーブル一覧(全26)",
    ["テーブル名", "概要", "主キー", "主要外部キー", "Unique", "主な利用機能", "RLS有無", "pgvector",
     "実装根拠(migration)", "備考"],
    TBL, [22, 30, 22, 36, 22, 18, 8, 12, 18, 18])

# ─────────────────────────── 11_DB項目定義 ───────────────────────────
COLS = [
    ["profiles", "id", "uuid", "NOT NULL", "—", "Yes", "—", "Yes", "—", "識別子", "認証ユーザーID", "0001"],
    ["profiles", "role", "user_role(enum)", "NOT NULL", "—", "—", "—", "—", "—", "—", "job_seeker/company_member", "0001"],
    ["profiles", "display_name", "text", "NOT NULL", "—", "—", "—", "—", "—", "直接PII", "表示名", "0001"],
    ["profiles", "locale", "locale_code(enum)", "NOT NULL", "'ja'", "—", "—", "—", "—", "—", "言語設定", "0001"],
    ["candidates", "user_id", "uuid", "NOT NULL", "—", "—", "profiles(id)", "Yes", "—", "—", "所有者", "0002"],
    ["candidates", "headline/summary/location", "text", "NULL", "—", "—", "—", "—", "—", "準PII", "プロフィール本文", "0002"],
    ["candidates", "years_experience", "numeric(4,1)", "NOT NULL", "0", "—", "—", "—", "—", "—", "総経験年数", "0002"],
    ["candidates", "open_to_work", "boolean", "NOT NULL", "true", "—", "—", "—", "—", "—", "公開可否(可視性)", "0002"],
    ["candidates", "desired_salary_min/max", "integer", "NULL", "—", "—", "—", "—", "—", "PII", "希望給与", "0002"],
    ["resume_files", "storage_path", "text", "NOT NULL", "—", "—", "—", "Yes", "—", "—", "推測不可のサーバ生成パス", "0002"],
    ["resume_files", "scan_result", "scan_result(enum)", "NOT NULL", "'clean'", "—", "—", "—", "—", "—", "clean/infected/error", "0002"],
    ["resume_files", "extracted_text", "text", "NULL", "—", "—", "—", "—", "—", "PII", "抽出本文", "0002"],
    ["jobs", "status", "job_status(enum)", "NOT NULL", "'draft'", "—", "—", "—", "—", "—", "draft/open/closed", "0004"],
    ["jobs", "visibility", "job_visibility(enum)", "NOT NULL", "'private'", "—", "—", "—", "—", "—", "public/private", "0004"],
    ["jobs", "salary_min/max", "integer", "NULL", "—", "—", "—", "—", "max≧min", "—", "給与範囲", "0004"],
    ["match_results", "score", "integer", "NOT NULL", "—", "—", "—", "—", "0..100", "—", "マッチスコア", "0006"],
    ["match_results", "breakdown", "jsonb", "NOT NULL", "—", "—", "—", "—", "—", "—", "内訳(ベクトル/ルール)", "0006"],
    ["candidate_embeddings", "embedding", "vector(384)", "NOT NULL", "—", "—", "—", "—", "—", "—", "L2正規化埋め込み", "0005"],
    ["messages", "body", "text", "NOT NULL", "—", "—", "—", "—", "—", "PII", "メッセージ本文", "0007"],
    ["messages", "client_token", "uuid", "NULL", "—", "—", "—", "(conv,sender,token)", "—", "—", "重複排除キー", "0007"],
    ["applications", "stage", "recruitment_stage(enum)", "NOT NULL", "'applied'", "—", "—", "—", "—", "—", "応募段階", "0008"],
    ["interviews", "status", "interview_status(enum)", "NOT NULL", "'proposed'", "—", "—", "—", "—", "—", "面接状態", "0008"],
    ["job_queue", "idempotency_key", "text", "NULL", "—", "—", "—", "Yes", "—", "—", "冪等キー", "0011"],
]
add(wb, "11_DB項目定義", "11_DB項目定義(主要/個人情報テーブル)",
    ["テーブル名", "カラム名", "データ型", "Null可否", "デフォルト", "主キー", "外部キー", "Unique",
     "Check制約", "個人情報区分", "説明", "実装根拠"],
    COLS, [20, 24, 20, 12, 10, 8, 14, 16, 10, 12, 28, 10])

# ─────────────────────────── 12_ERD_RLS概要 ───────────────────────────
add(wb, "12_ERD_RLS概要", "12_ERD・RLS概要(主要関係)",
    ["領域", "エンティティ/関係", "From", "To", "関係種別", "多重度", "説明", "RLS観点", "実装根拠"],
    [
        ["認証", "所有", "profiles", "candidates", "1:1", "1対1", "ユーザーは1つの候補者を持つ", "owns_candidate", "0002,0009"],
        ["履歴書", "保有", "candidates", "resume_files", "1:N", "1対多", "候補者は複数履歴書を持つ", "owns_candidate", "0002,0009"],
        ["分析", "生成", "candidates", "skill_analyses", "1:N", "1対多", "候補者の分析履歴", "candidate_readable", "0002,0009"],
        ["企業", "所属", "profiles", "company_members", "1:N", "1対多", "ユーザーは複数企業に所属可", "is_company_member", "0003,0009"],
        ["求人", "公開", "companies", "jobs", "1:N", "1対多", "企業は複数求人を持つ", "job_readable/is_company_member", "0004,0009"],
        ["マッチ", "算定", "jobs×candidates", "match_results", "N:N", "多対多", "求人×候補者のスコア", "owns/ is_company_member", "0006,0009"],
        ["応募", "応募", "candidates", "applications", "1:N", "1対多", "候補者の応募", "application_readable", "0008,0009"],
        ["選考", "履歴", "applications", "application_stage_history", "1:N", "1対多", "段階遷移の監査", "application_readable", "0008,0009"],
        ["面接", "提案", "applications", "interviews", "1:N", "1対多", "応募ごとの面接", "application_readable", "0008,0009"],
        ["会話", "参加", "conversations", "conversation_members", "1:N", "1対多", "スレッド参加者", "is_conversation_member", "0007,0009"],
        ["通知", "宛先", "profiles", "notifications", "1:N", "1対多", "ユーザー宛通知", "user_id=auth.uid()", "0007,0009"],
    ],
    [10, 20, 18, 24, 10, 8, 28, 28, 14])

# ─────────────────────────── 13_RLS権限マトリクス ───────────────────────────
RLS = [
    ["profiles", "SELECT", "×", "○", "○", "○", "true(全認証可視)", "—", "—", "0009", "rls.rls.test.ts"],
    ["profiles", "INSERT/UPDATE", "×", "○(自)", "○(自)", "○", "id=auth.uid()", "自IDのみ", "自己のみ更新", "0009", "rls.rls.test.ts"],
    ["candidates", "SELECT", "×", "○(条件)", "○(条件)", "○", "candidate_readable(id)", "公開/応募/保存のみ", "可視性関数で限定", "0009", "security-negative.rls.test.ts"],
    ["candidates", "INSERT/UPDATE", "×", "○(自)", "×", "○", "user_id=auth.uid()", "本人のみ", "他者不可", "0009", "rls.rls.test.ts"],
    ["resume_files", "ALL", "×", "○(自)", "×", "○", "owns_candidate", "本人のみ", "他者の履歴書不可", "0009", "security-files.integration.test.ts"],
    ["skill_analyses", "INSERT/UPDATE", "×", "×", "×", "○", "serviceのみ", "—", "信頼経路のみ書込", "0009", "—"],
    ["companies", "INSERT", "×", "○(role)", "○(role)", "○", "viewer_is_company_role()", "—", "—", "0009", "rls.rls.test.ts"],
    ["companies", "UPDATE/DELETE", "×", "×", "○(member)", "○", "is_company_member(id)", "企業境界", "他社編集不可", "0009", "security-negative.rls.test.ts"],
    ["jobs", "SELECT", "○(open+public)", "○", "○", "○", "(open∧public)∨member", "—", "下書き/非公開は秘匿", "0009", "company-job.integration.test.ts"],
    ["jobs", "INSERT/UPDATE/DELETE", "×", "×", "○(member)", "○", "is_company_member(company_id)", "企業境界", "他社求人不可", "0009", "security-negative.rls.test.ts"],
    ["match_results", "SELECT", "×", "○(自)", "○(自社求人)", "○", "owns∨is_company_member(job)", "両者限定", "他者スコア不可", "0009", "matching.integration.test.ts"],
    ["messages", "INSERT", "×", "○(member)", "○(member)", "○", "sender=auth.uid()∧member", "会話境界", "非メンバ投稿不可", "0009", "messaging.integration.test.ts"],
    ["notifications", "SELECT/UPDATE", "×", "○(自)", "○(自)", "○", "user_id=auth.uid()", "本人のみ", "他者通知不可", "0009", "messaging.integration.test.ts"],
    ["applications", "SELECT", "×", "○(自)", "○(自社求人)", "○", "owns∨is_company_member(job)", "両者限定", "他者応募不可", "0009", "recruitment.integration.test.ts"],
    ["shortlists/comparisons", "ALL", "×", "×", "○(member)", "○", "is_company_member(company)", "企業境界", "他社不可", "0009", "recruitment.integration.test.ts"],
    ["interviews", "INSERT", "×", "×", "○(member)", "○", "proposed_by=auth.uid()∧member", "企業境界", "—", "0009", "recruitment.integration.test.ts"],
    ["Storage:resumes", "DOWNLOAD", "×", "○(owner)", "API仲介", "○", "owner=auth.uid()", "所有者パス分離", "企業はAPI経由のみ", "0010", "security-files.integration.test.ts"],
    ["job_queue", "ALL", "×", "×", "×", "○", "serviceのみ", "—", "内部ワーカー専用", "0011", "jobqueue.integration.test.ts"],
]
add(wb, "13_RLS権限マトリクス", "13_RLS権限マトリクス",
    ["テーブル/Storage", "操作", "anonymous", "seeker", "company_member", "admin/system",
     "条件(USING/WITH CHECK)", "テナント分離ルール", "IDOR対策", "実装根拠", "テスト根拠"],
    RLS, [20, 16, 12, 12, 16, 12, 26, 18, 18, 10, 26])

# ─────────────────────────── 14_AI_マッチング設計 ───────────────────────────
add(wb, "14_AI_マッチング設計", "14_AI・マッチング設計",
    ["設計項目", "処理種別", "入力", "出力", "Provider/Adapter", "Deterministic mock", "Prompt/Schema対策",
     "Embedding", "Score/Algorithm", "Failure handling", "実装根拠", "テスト根拠"],
    [
        ["スキル分析", "LLM抽出+JSON生成", "履歴書本文(≤20k)+ロケール", "SkillAnalysisResult(スキル/年数/提案/学習)",
         "mock / 外部LLM(Anthropic) / 外部LLM(OpenAI)(設定で切替)", "MockLlmProvider(辞書走査・外部呼出なし)",
         "wrapUntrustedDocument(区切り+制御文字除去)+Zod検証", "—", "—",
         "3回再試行+回路遮断+検証失敗時は前回有効値", "apps/api/src/services/analysis.ts, adapters/ai/*",
         "providers.unit.test.ts, matching.integration.test.ts"],
        ["埋め込み生成", "テキスト→384次元", "候補者/求人の特徴テキスト", "vector(384) L2正規化",
         "mock / openai(text-embedding-3-small,384)", "MockEmbeddingProvider(SHA1ハッシュ語袋・決定論的)",
         "入力は上流で無害化済", "384次元・cosine", "—", "回路遮断(15s/2再試行)+source_hash不変ならスキップ",
         "apps/api/src/adapters/ai/embedding.ts, services/embeddings.ts", "scoring.test.ts"],
        ["マッチング算定", "決定論的ルール合成", "候補者・求人の構造化属性+ベクトル類似度",
         "score(0..100)+breakdown+reason+一致/不足スキル", "scoreMatch(純粋関数)", "✓(同入力→同出力)",
         "—(構造化データ)", "vector0.30+skill0.35+exp0.15+sal0.08+loc0.07+lang0.05(match-v1)",
         "純粋関数で失敗経路なし。algorithm_versionで再現保証", "packages/shared/src/scoring.ts",
         "scoring.test.ts(12), matching.integration.test.ts(再現性)"],
        ["ベクトル近傍探索", "pgvector k-NN", "候補者/求人ID(埋め込み)", "上位50近傍(cosine)",
         "pgvector ivfflat(<=>)", "—(DBネイティブ)", "—", "384次元・lists=10・probes=10",
         "cosine類似度=1-距離", "埋め込み欠如はrecall=0扱い", "apps/api/src/services/matching.ts", "matching.integration.test.ts"],
        ["プロンプト安全", "サニタイズ+秘匿", "未信頼文書/ログ文字列", "区切り付き文書/[redacted]済テキスト",
         "—(共有関数)", "—", "制御文字除去・区切りエスケープ・PII(メール/電話/長トークン)マスク", "—", "—",
         "—", "packages/shared/src/prompt-safety.ts", "prompt-safety.test.ts(7)"],
        ["障害耐性", "回路遮断+タイムアウト+再試行", "プロバイダ呼出", "結果またはupstreamAiError",
         "resilientLlm/resilientEmbedding", "—", "—", "—", "—",
         "閾値5回でopen・30s冷却・指数バックオフ・指標記録", "apps/api/src/resilience.ts, adapters/ai/resilient.ts",
         "resilience.unit.test.ts(6)"],
    ],
    [16, 16, 24, 26, 22, 26, 28, 18, 30, 26, 30, 26])

# ─────────────────────────── 15_バッチ_ジョブ設計 ───────────────────────────
add(wb, "15_バッチ_ジョブ設計", "15_バッチ・ジョブ設計",
    ["ジョブID", "ジョブ名", "起動契機", "入力", "処理内容", "冪等性", "Retry/Timeout", "Dead-letter/Failure",
     "監視", "実装状態", "実装根拠", "備考"],
    [
        ["resume_parse", "履歴書解析", "履歴書アップロード時にenqueue(再試行は手動API)", "parseJobId",
         "テキスト抽出→AI分析→スキル/年数保存→候補者埋め込み再生成", "idempotency_key(parseJobId)+状態遷移で再処理防止",
         "最大5回・指数バックオフ(上限300s)・ジョブ30s timeout", "上限超過でstatus=dead・last_error記録",
         "job_queue状態+指標(job_resume_parse_duration, jobs_processed_total)",
         "local/test=インライン即時drain / 本番=常駐ワーカー(FOR UPDATE SKIP LOCKED)",
         "apps/api/src/jobs/{queue,handlers}.ts, services/resume.ts", "起動時start()/SIGTERMでstop()ドレイン"],
    ],
    [12, 14, 30, 12, 34, 28, 26, 24, 28, 30, 28, 22])

# ─────────────────────────── 16_多言語設計 ───────────────────────────
add(wb, "16_多言語設計", "16_多言語設計",
    ["Locale", "対象範囲", "i18n実装方式", "主な辞書ファイル", "APIエラー対応", "未翻訳検知", "テスト根拠", "備考"],
    [
        ["ja", "UI全体(既定)", "react-i18next(即時ロード)", "i18n/locales/ja.ts",
         "messageKey→辞書解決(localizeError)", "パリティテストでキー整合", "i18n.parity.test.ts", "DEFAULT_LOCALE。fallback先頭"],
        ["en", "UI全体(正規)", "react-i18next(即時ロード)", "i18n/locales/en.ts",
         "同上(英語fallback)", "enを正準としキー集合を比較", "i18n.parity.test.ts", "キー定義の基準"],
        ["zh-CN", "UI全体", "react-i18next(コード分割・遅延)", "i18n/locales/zh-CN.ts",
         "同上", "葉(非空文字列)とキーを検証", "i18n.parity.test.ts", "切替時にensureBundleで動的import"],
        ["zh-TW", "UI全体", "react-i18next(コード分割・遅延)", "i18n/locales/zh-TW.ts",
         "同上", "同上", "i18n.parity.test.ts", "同上"],
    ],
    [10, 18, 28, 24, 28, 22, 22, 30])

# ─────────────────────────── 17_非機能_運用設計 ───────────────────────────
add(wb, "17_非機能_運用設計", "17_非機能・運用設計",
    ["分類", "設計項目", "要求内容", "現行実装", "評価", "証跡", "リスク", "改善案"],
    [
        ["セキュリティ", "HTTPヘッダ", "安全なヘッダを付与する", "helmetで厳格CSP/CORP/no-referrer/HSTS(本番)", "一致",
         "apps/api/src/app.ts", "CSP調整漏れ", "デプロイ毎にヘッダ検査"],
        ["セキュリティ", "認可", "ロールとテナントを分離する", "RLS+SET LOCAL ROLE+SECURITY DEFINER関数", "一致",
         "supabase/migrations/0009_rls.sql", "ポリシー追加漏れ", "RLS否定テストを継続"],
        ["セキュリティ", "レート制限", "濫用を抑止する", "全体+経路別(登録20/アップロード10/送信60/採用30)", "一致",
         "app.ts, routes/*", "閾値調整", "監視と動的調整"],
        ["プライバシー", "PII保護", "個人情報を保護する", "ログredact+prompt-safety+所有者限定RLS", "一致",
         "logger.ts, prompt-safety.ts", "新規PII列の見落し", "PII区分の定期棚卸"],
        ["監査", "段階履歴", "選考の変更を追跡する", "application_stage_historyに記録", "一致",
         "0008, recruitment.ts", "—", "—"],
        ["ログ", "構造化ログ", "秘匿しつつ追跡する", "pino+相関ID(x-correlation-id)+秘匿", "一致",
         "app.ts, logger.ts", "—", "集約基盤連携"],
        ["可観測性", "計測/健全性", "稼働を可視化する", "/health /ready /metrics(Prometheus)", "実装のみ(要求書未記載)",
         "app.ts, observability/metrics.ts", "メトリクス未集約", "OTel/Prometheus導入"],
        ["性能", "予算/計測", "応答性能を担保する", "バンドル予算142KB<200KB・Lighthouse・p50/p95計測", "一致",
         "scripts/{check-bundle,check-lighthouse}.mjs, PERFORMANCE.md", "本番値未測定", "supabaseで実測"],
        ["可用性", "信頼性", "外部障害に耐える", "回路遮断+timeout+再試行+耐久キュー+graceful shutdown", "実装のみ(要求書未記載)",
         "resilience.ts, jobs/queue.ts, server.ts", "—", "—"],
        ["アクセシビリティ", "WCAG", "操作性を確保する", "axeで重大/深刻0・キーボード操作・Lighthouse a11y 100", "一致",
         "apps/web/e2e/a11y.spec.ts", "新規画面の退行", "CIでa11yゲート継続"],
        ["デプロイ", "コンテナ", "再現可能に配備する", "多段非rootのDockerfile+compose+CI", "実装のみ(要求書未記載)",
         "infra/docker/*, .github/workflows/ci.yml", "本番デプロイ未実施", "段階導入"],
        ["バックアップ", "データ保護", "消失に備える", "本番DBはSupabase(PITR想定)・手順を文書化", "確認不能",
         "docs/OPERATIONS.md", "未設定リスク", "PITR有効化を確認"],
        ["移行", "マイグレーション", "前進専用で安全に移行する", "0001〜0012・追加専用・migrate:check検証", "一致",
         "supabase/migrations, db/migrate-check.ts", "本番適用は手動", "適用手順をrunbook化"],
        ["ファイルアップロード", "検証", "不正ファイルを排除する", "10MB上限・拡張子/MIME/マジックバイト整合", "一致",
         "services/resume.ts, validateUpload", "—", "—"],
        ["ウイルス対策", "スキャン", "感染ファイルを保存しない", "本番clamav(モック禁止)・local mock", "一致",
         "config.ts(起動時拒否), services/resume.ts", "clamd未配備", "本番でclamd配備確認"],
        ["データ保持", "削除/保持", "保持と削除を運用する", "運用者手順を文書化(自助削除UIは無し)", "確認不能",
         "docs/OPERATIONS.md", "要件未確定", "要件確定後に手順整備"],
    ],
    [14, 16, 26, 34, 18, 30, 18, 22])

# ─────────────────────────── 18_テスト設計 ───────────────────────────
add(wb, "18_テスト設計", "18_テスト設計",
    ["テスト分類", "対象機能", "テスト観点", "テストファイル", "主なケース", "件数", "実行コマンド", "結果", "備考"],
    [
        ["unit(共有)", "スコアリング/検証/プロンプト安全", "決定論・境界・PII", "packages/shared/src/*.test.ts",
         "score境界/再現性, Zod, redact", "35", "npm run test -w @ars/shared", "35 passed", "3ファイル"],
        ["unit(API)", "プロバイダ/障害耐性/ロガー", "選択・遮断・秘匿", "apps/api/src/*.unit.test.ts",
         "mock決定論, 回路遮断, redact", "(api計121に含む)", "npm run test:unit -w @ars/api", "passed", "—"],
        ["integration(API)", "認証/履歴書/解析/企業求人/マッチ/人材/メッセージ/採用/契約/レート/冪等/健全性",
         "apps/api/tests/*.integration.test.ts", "正常+異常+権限+境界", "78", "npm run test:integration -w @ars/api",
         "78 passed", "OpenAPI契約・429・並行応募含む"],
        ["rls(DB)", "行レベルセキュリティ(肯定/否定)", "可視性・テナント分離・IDOR", "apps/api/tests/*.rls.test.ts",
         "本人のみ可視, 他社不可, 匿名不可", "26", "npm run test:rls -w @ars/api", "26 passed", "—"],
        ["component(Web)", "デザイン/フォーム/フック/ページ", "描画・操作・i18n", "apps/web/src/**/*.test.tsx",
         "Button/Field/Dialog, ドロップゾーン, 言語切替, 離脱警告, パリティ", "25", "npm run test -w @ars/web",
         "25 passed", "8ファイル"],
        ["e2e", "主要ジャーニー(求職者/企業)", "登録→分析→応募/求人→選考", "apps/web/e2e/journeys.spec.ts",
         "2ロールの一気通貫", "4", "npm run test:e2e", "4 passed", "Playwright"],
        ["a11y", "アクセシビリティ", "重大/深刻0・キーボード", "apps/web/e2e/a11y.spec.ts",
         "公開/状態/認証画面のaxe走査+キーボード", "3", "npm run test:a11y", "3 passed", "WCAG2.1 AA"],
        ["visual", "視覚回帰/品質", "溢れ・コンソールエラー0", "apps/web/e2e/visual.spec.ts",
         "5幅×2言語×11画面=110枚", "1", "npm run test:visual", "1 passed", "スクリーンショット110枚"],
        ["合計", "—", "—", "—", "—", "189(vitest181+Playwright8)", "npm run verify ほか", "全て成功", "計測commit 1a46f42"],
    ],
    [14, 30, 18, 30, 30, 18, 28, 12, 22])

# ─────────────────────────── 19_トレーサビリティ ───────────────────────────
TRC_NOTE = "要求仕様(FR-01〜FR-10)を基準とする。要求仕様書の細目と未突合の事項は要確認とする。"
add(wb, "19_トレーサビリティ", "19_トレーサビリティ(FR-01〜FR-10)",
    ["要件ID", "要件概要", "設計シート", "画面ID", "API ID", "DB/RLS", "テスト", "実装状態", "差異分類", "証跡", "備考"],
    [
        ["FR-01", "認証・ユーザー管理", "03,04,06,08", "SC-02/03/04/09", "A-01〜A-05",
         "profiles / profiles_*", "auth.integration.test.ts", "実装済", "一致", "routes/auth.ts", TRC_NOTE],
        ["FR-02", "履歴書アップロード・解析", "04,07,14,15", "SC-12", "A-08〜A-11,A-14",
         "resume_files,parse_jobs / owns", "resume.integration.test.ts, security-files", "実装済", "一致",
         "services/resume.ts", "マジックバイト+ウイルス検査"],
        ["FR-03", "AIスキル分析", "04,14", "SC-12", "A-12,A-13",
         "skill_analyses,candidate_skills", "matching.integration.test.ts", "実装済", "一致",
         "services/analysis.ts", "Zod検証+再試行"],
        ["FR-04", "企業・求人管理", "04,06,08", "SC-15/16/17", "A-15,A-19,A-24〜A-27",
         "companies,jobs / is_company_member", "company-job.integration.test.ts", "実装済", "一致",
         "routes/{company,job}.ts", "公開範囲制御"],
        ["FR-05", "AIマッチング", "04,14", "SC-13", "A-28,A-29",
         "match_results,*_embeddings / owns,member", "matching.integration.test.ts, scoring.test.ts", "実装済", "一致",
         "services/matching.ts, scoring.ts", "決定論・版管理"],
        ["FR-06", "人材検索・候補者詳細", "04,06,08", "SC-18/19", "A-30,A-31,A-42",
         "candidates / candidate_readable", "talent.integration.test.ts", "実装済", "一致",
         "services/talent.ts", "連絡先/履歴書は関係成立時のみ"],
        ["FR-07", "企業・求人検索", "04,06,08", "SC-05/06/07/08", "A-16,A-18,A-21,A-22,A-23",
         "jobs,companies / 公開select", "company-job.integration.test.ts", "実装済", "一致",
         "routes/{job,company}.ts", "anonはopen+publicのみ"],
        ["FR-08", "メッセージ・通知", "04,06,08", "SC-10/11", "A-32〜A-38",
         "conversations,messages,notifications", "messaging.integration.test.ts", "実装済", "一致",
         "routes/messaging.ts", "client_token重複排除"],
        ["FR-09", "多言語対応(4言語)", "16", "全画面", "—(messageKey)",
         "—(クライアント辞書)", "i18n.parity.test.ts", "実装済", "一致",
         "apps/web/src/i18n/*", "zh-*遅延ロード"],
        ["FR-10", "比較・面接・採用フロー", "04,05,08", "SC-14/17/18", "A-39〜A-43",
         "applications,interviews,shortlists / readable", "recruitment.integration.test.ts", "実装済", "一致",
         "routes/recruitment.ts", "段階遷移権限"],
        ["FR-02.4", "ウイルススキャン必須", "17", "SC-12", "A-08", "resume_files.scan_result",
         "security-files.integration.test.ts", "実装済", "一致", "config.ts(本番でmock禁止)", "—"],
        ["FR-09.4", "言語切替の永続化", "16", "全画面", "—", "localStorage(ars.locale)",
         "LanguageSwitcher.test.tsx", "実装済", "一致", "i18n/index.ts", "—"],
        ["要求細目", "要求仕様書の個別要求文言", "20,21", "—", "—", "—", "—", "—", "確認不能",
         "要件担当と確認", "FR列挙以上の細目は要件担当と突合する"],
    ],
    [10, 24, 12, 16, 22, 26, 28, 10, 14, 22, 30])

# ─────────────────────────── 20_差異_未決事項 ───────────────────────────
add(wb, "20_差異_未決事項", "20_差異・未決事項",
    ["No.", "区分", "差異分類", "内容", "要件側", "実装側", "業務影響", "技術影響", "推奨対応", "優先度", "確認先", "証跡"],
    [
        ["1", "要求突合", "確認不能", "要求仕様書の細目とFR単位を超えた突合が未完了である", "要求仕様書(FR-01〜FR-10)",
         "実装と整合", "要求細目の取りこぼし可能性", "—", "要件担当と細目を突合する", "高",
         "要件担当", "00,02,21"],
        ["2", "可観測性", "実装のみ（要求書未記載）", "/metricsとPrometheus露出, 相関ID", "記載未確認",
         "app.ts, metrics.ts", "運用容易化", "監視基盤前提", "要求への正式取込", "中", "運用", "17"],
        ["3", "信頼性", "実装のみ（要求書未記載）", "耐久ジョブキュー・回路遮断・graceful shutdown", "記載未確認",
         "jobs/queue.ts, resilience.ts", "障害耐性向上", "運用設計増", "要求への正式取込", "中", "運用", "14,15"],
        ["4", "セキュリティ", "実装のみ（要求書未記載）", "レート制限の経路別強化・SBOM・依存監査", "記載未確認",
         "routes/*, scripts/*", "濫用/供給網対策", "—", "ポリシー文書化", "中", "セキュリティ", "17"],
        ["5", "データ保持", "確認不能", "保持期間・削除要求対応の要件が未確定である", "要件未確定",
         "運用者手順のみ(自助UI無)", "コンプライアンス", "削除はFK順序考慮要", "保持方針を要件確定する", "高",
         "法務/要件担当", "17, docs/OPERATIONS.md"],
        ["6", "アカウント削除", "部分実装", "自助の退会/削除UIは存在せず運用者手動である", "未確認",
         "PATCH /auth/accountのみ", "退会導線なし", "actor系FKはRESTRICT", "要件に応じ削除APIを検討", "中",
         "要件", "docs/DATABASE.md, OPERATIONS.md"],
        ["7", "本番計測", "確認不能", "本番相当環境の実測値が未取得である", "要件未確定",
         "検証環境の実測値のみ記録", "性能目標の確定不可", "—", "本番相当環境で再計測する", "中", "運用", "17, PERFORMANCE.md"],
    ],
    [5, 16, 22, 30, 18, 26, 20, 18, 28, 8, 14, 22])

# ─────────────────────────── 21_証跡 ───────────────────────────
add(wb, "21_証跡", "21_証跡(調査対象一覧)",
    ["証跡ID", "種別", "パス", "内容要約", "関連シート", "確認結果", "備考"],
    [
        ["E-01", "要求資料", "AIを実装した人材採用マッチングシステム開発_仕様説明.docx", "要求仕様書", "00,19,20",
         "要確認", "要件担当と細目を突合予定"],
        ["E-02", "要求範囲", "要求仕様(FR-01〜FR-10)", "機能要求の識別子定義", "02,04,19", "確認", "要求範囲の根拠"],
        ["E-03", "フロント", "apps/web/src/{App.tsx,pages,components,design,i18n,hooks,lib}", "SPA・23画面・i18n", "06,07,16", "確認", "—"],
        ["E-04", "ルート", "apps/api/src/routes/*.ts", "全APIルート(約44)", "08,09", "確認", "auth/candidate/company/job/match/talent/messaging/recruitment"],
        ["E-05", "サービス", "apps/api/src/services/*.ts", "ドメインロジック", "08,14,15", "確認", "—"],
        ["E-06", "共有", "packages/shared/src/*", "Zodスキーマ・スコアリング・プロンプト安全", "07,09,14", "確認", "—"],
        ["E-07", "マイグレーション", "supabase/migrations/0001〜0012", "26テーブル・型・索引・pgvector", "10,11,12", "確認", "前進専用"],
        ["E-08", "RLS", "supabase/migrations/0009_rls.sql", "行レベルセキュリティ・定義関数", "12,13", "確認", "—"],
        ["E-09", "ストレージ", "supabase/migrations/0010_storage.sql", "resumesバケット・ポリシー", "13", "確認", "所有者パス分離"],
        ["E-10", "AIアダプタ", "apps/api/src/adapters/ai/*, resilience.ts", "LLM/埋め込み/障害耐性", "14", "確認", "mock/anthropic/openai"],
        ["E-11", "ジョブ", "apps/api/src/jobs/*", "耐久キュー・ハンドラ", "15", "確認", "resume_parse"],
        ["E-12", "i18n辞書", "apps/web/src/i18n/locales/{ja,en,zh-CN,zh-TW}.ts", "4言語辞書", "16", "確認", "パリティテスト有"],
        ["E-13", "テスト", "apps/api/tests, packages/shared/src/*.test.ts, apps/web/src/**/*.test.tsx, apps/web/e2e/*.spec.ts",
         "unit/integration/rls/component/e2e/a11y/visual", "18", "確認", "計189(commit 1a46f42)"],
        ["E-14", "既存ドキュメント", "docs/{ARCHITECTURE,API,DATABASE,AI,DEPLOY,OPERATIONS,RUNBOOK,SECURITY,PERFORMANCE}.md",
         "設計・運用・性能の既存文書", "03,17", "確認", "9文書"],
        ["E-15", "CI/インフラ", ".github/workflows/ci.yml, infra/docker/*", "品質ゲート・コンテナ", "17", "確認", "変更せず参照のみ"],
    ],
    [8, 16, 44, 28, 14, 16, 30])

# ─────────────────────────── 99_用語集 ───────────────────────────
add(wb, "99_用語集", "99_用語集",
    ["用語", "区分", "説明", "関連", "備考"],
    [
        ["RLS", "DB", "行レベルセキュリティ。ロール/所有/テナントで行アクセスを制御する。", "13", "PostgreSQL機能"],
        ["pgvector", "AI/DB", "ベクトル型と近傍探索を提供する拡張。ivfflatでcosine探索する。", "14", "384次元"],
        ["埋め込み(embedding)", "AI", "テキストを固定長ベクトルへ写像したもの。", "14", "L2正規化"],
        ["match-v1", "AI", "マッチング算定の版。重みと算定式を固定し再現性を担保する。", "14", "algorithm_versions"],
        ["ランタイム(ARS_RUNTIME)", "基盤", "local(PGlite+モック)とsupabase(本番)の二系統である。", "03", "同一コード"],
        ["job_seeker/company_member", "認可", "求職者/企業担当者のロールである。", "13", "profiles.role"],
        ["冪等性(idempotency)", "ジョブ", "同一鍵で重複投入しても一度だけ処理される性質である。", "15", "job_queue"],
        ["である調", "文書", "敬体を用いない常体の文体である。本書で一貫採用する。", "02", "—"],
        ["FR-01〜FR-10", "要求", "本システムの機能要求の識別子である。", "04,19", "要求範囲"],
    ],
    [22, 12, 50, 12, 18])

wb.save(OUT)
import os
size = os.path.getsize(OUT)
print("SHEETS:", len(wb.sheetnames))
print("NAMES:", wb.sheetnames)
print(f"SAVED: {OUT} ({size} bytes)")

